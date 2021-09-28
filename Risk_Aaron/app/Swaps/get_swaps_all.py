import requests
from bs4 import BeautifulSoup

import datetime
from app.decorators import async_fun

now = datetime.datetime.now()

from io import StringIO
from bs4 import BeautifulSoup

from sqlalchemy import create_engine, text

from requests.auth import HTTPBasicAuth  # or HTTPDigestAuth, or OAuth1, etc.
from requests import Session
from zeep.transports import Transport
from suds.client import Client
import logging
from zeep import Client
from zeep import helpers

import pandas as pd
import numpy as np

from Aaron_Lib import *

from unsync import unsync

from app.OZ_Rest_Class import *

from Helper_Flask_Lib import *
from flask import flash
from flask_login import current_user

from sklearn.linear_model import LinearRegression

from sqlalchemy import text

# For Excel saving
import pyexcel as pe

# Openpyxl for excel styling.
# Different lib from pyexcel
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


# For moving Files around.
from os import listdir
from os.path import isfile, join


#logging.basicConfig(level=logging.INFO)
#logging.getLogger('suds.client').setLevel(logging.INFO)


# From https://www.fxdd.com/mt/en/trading/offering
# FXdd Pricing is from https://www.fxdd.com/api/price-feed/clear as Json
# Price_URL_Clear = "https://www.fxdd.com/api/price-feed/clear"
# Price_URL_Standard = "https://www.fxdd.com/api/price-feed/standard"



@unsync
def get_swaps_fxdd():
    # https://www.fxdd.com/mt/en/trading/offering
    try:
        # url = "http://www.fxdd.com/fileadmin/setup/libraries/proxy/jason.php"   #The URL for their JSON
        url = "https://secure.fxdd.com/fileadmin/setup/libraries/proxy/jason.php"  # The URL that return JSON. Changed endpoint on 21 Nov 2019

        r = requests.get(url)

        if r.status_code != 200:  # If request failed
            return {}

        data = r.text  # Get the text form
        Fxdd_Swap_Json_String = StringIO(r.text)
        fxdd_Swap_Json = json.load(Fxdd_Swap_Json_String)

        if len(fxdd_Swap_Json) < 2:  # If length is less than 2
            return []

        # Want to get the index of "buy" and "sell" programmably in case they swap it around.
        long_index = max(
            [i if fxdd_Swap_Json[0][i].lower().find("buy") >= 0 else -1 for i in range(len(fxdd_Swap_Json[0]))])
        short_index = max(
            [i if fxdd_Swap_Json[0][i].lower().find("sell") >= 0 else -1 for i in range(len(fxdd_Swap_Json[0]))])
        symbol_index = max(
            [i if fxdd_Swap_Json[0][i].lower().find("cp") >= 0 else -1 for i in range(len(fxdd_Swap_Json[0]))])

        if long_index == -1 or short_index == -1 or symbol_index == -1:  # Return error.
            return []

        # Want to get the long and short in float, IF it is, else, don't want it.
        # also, to replace "/"
        return_list_dict = [
            {"Symbol": s[symbol_index].replace("/", ""), "Long": float(s[long_index]), "Short": float(s[short_index])}
            for s in fxdd_Swap_Json[1:] if isfloat(s[long_index]) and isfloat(s[short_index])]

        # for s in fxdd_Swap_Json[1:]:    # Since the first one is the column names
        #     if isfloat(s[long_index]) and isfloat(s[short_index]):
        #         multiplier = 100000     # Not given in points. Need to multiply by digits
        #         if s[symbol_index][-3:].find("JPY") >= 0:
        #             multiplier = 1000
        #         elif s[symbol_index][0].find("X") >= 0:
        #             multiplier = 100
        #
        #
        #         return_list_dict.append({"Symbol": s[symbol_index].replace("/",""),
        #                                  "Long" : -1*multiplier*float(s[long_index]), "Short": multiplier*float(s[short_index])})

        return return_list_dict
    except:
        return []

@unsync
def get_swaps_forexDotCom():
    # Information got off https://www.forex.com/en/trading/pricing-fees/rollover-rates/

    # # The Original payload header.
    # payload = {"siteId": "forex.web.g2en", "products": [], "marketType": "FX", "productType": "FX",
    #  "requiredFields": ["Product", "Long", "Short", "AliasName", "Product", "SEOName", "MarketId",
    #                     "HasMultipleProductTypes"]}

    # Forex.com (Rollover rates displayed are based on a 10K position)
    # We need 1 lot. So, multiply by 10
    multiplier = 10

    # Has CFD. But we will not do the conversion as we it dosn't show on the website.

    # CHN.50
    # USA.30
    # ESP.35
    # FRA.40
    # GER.30
    # HKG.33
    # JPN225
    # NAS100
    # OILUSD
    # SPX500
    # UK.100

    try:
        # We took off the fields that we do not need
        payload = {"siteId": "forex.web.g2en", "products": [], "marketType": "FX", "productType": "FX",
                   "requiredFields": ["Product", "Long", "Short"]}

        url = "https://www.forex.com/_Srvc/feeds/LiveRates.asmx/GetMostPopularMarkets"

        return_data = requests.post(url, json=payload)

        if return_data.status_code != 200:
            print("Something went wrong with the Post request to forex.com while getting swaps.")
            return []

        swaps = return_data.json()

        if 'd' not in swaps:  # Usually swaps would be in swaps['d'] when it is returned from forex.com
            print("Something went wrong with the Post request to forex.com while getting swaps. No 'd' field")
            return []

        return_list_dict = []
        for d in swaps['d']:

            # Need to check if all the keys are in.
            if all([k in d for k in ["Product", "Long", "Short"]]) and isfloat(d['Long']) and isfloat(d['Short']):
                product = d['Product'] if 'Product' in d else ""
                long = round(multiplier * d['Long'],
                             2) if 'Long' in d else 0  # Need to multiply to our contract size, since fdc does 10K sizes
                short = round(multiplier * d['Short'],
                              2) if 'Short' in d else 0  # Need to multiply to our contract size, since fdc does 10K sizes
                swap_val_dict = {"Symbol": product.replace("/", ""), "Long": float(long), "Short": float(short)}
                return_list_dict.append(swap_val_dict)
        return return_list_dict
    except:
        return []


# Need to correct for digits
# Need to check if today isn't available.
# It's usually 1 day late.
@unsync
def get_swaps_saxo():
    try:
        # https://www.home.saxo/en-sg/rates-and-conditions/forex/trading-conditions#historic-swap-points

        payload = {"date": "{}".format(get_working_day_date(datetime.date.today(), -1).strftime("%Y-%m-%d"))}
        # If they don't have today's, need to check for yesterday, or the lsat working day

        return_data = requests.post("https://www.home.saxo/en-SG/rnc/hsp", json=payload)
        swaps = return_data.json()

        # Hardcode the dict keys as per json return.
        col_names = [j['name'] for j in swaps['rncHistoricSwapPoints']['tableData']['head'][1]['cells']]
        col_names[0] = "Symbol"  # Need to hard code since their table format is different.

        swap_val = [[i['cells'][j]['name'] for j in range(len(i['cells']))] for i in
                    swaps['rncHistoricSwapPoints']['tableData']['body']]

        swap_list_dict = [dict(zip(col_names, s)) for s in swap_val]

        # Saxo calculated 3 days swaps and upload triple values.
        # Could also use end-date - start-date.

        # We don't need the added information. Will only use the Long and Short
        swap_long_short = []

        # [{"Symbol": s['Symbol'], "Long": float(s["Long positions"]), "Short": float(s["Short positions"])}
        #                   for s in swap_list_dict if isfloat(s["Long positions"]) and isfloat(s["Short positions"])]

        for s in swap_list_dict:
            if isfloat(s["Long positions"]) and isfloat(s["Short positions"]):
                multiplier = 100000  # Not given in points. Need to multiply by digits
                if s['Symbol'][-3:].find("JPY") >= 0:  # JPY has lesser digits
                    multiplier = 1000
                elif s['Symbol'][0].find("X") >= 0:  # If it's PM, we need to multiply by 2 digits
                    multiplier = 100

                try:
                    # Need to try, since the string type might be different.
                    from_date = datetime.datetime.strptime(swap_list_dict[0]['From'], "%Y-%m-%d")
                    to_date = datetime.datetime.strptime(swap_list_dict[0]['To'], "%Y-%m-%d")
                    if abs((
                                   from_date - to_date).days) != 1:  # if it's one 1 day apart, we need to account for triple swaps
                        s["Long positions"] = float(s["Long positions"]) / 3
                        s["Short positions"] = float(s["Short positions"]) / 3
                except:
                    pass

                swap_long_short.append(
                    {"Symbol": s['Symbol'], "Long": round(-1 * multiplier * float(s["Long positions"]), 2),
                     "Short": round(multiplier * float(s["Short positions"]), 2)})

        return swap_long_short

    except:
        return []


# Scrape if off the website.
# Not by JSON
@unsync
def get_swaps_tradeview():
    try:
        url = "https://www.tradeviewforex.com/room/forex-resources/rollover-rates"

        return_data = requests.get(url=url)
        soup = BeautifulSoup(return_data.text, 'html.parser')
        table = soup.find_all("table")
        if len(table) == 0:
            return {}

        tradeview_to_bgi_symbols = {'WS30': '.US30', 'UK100': '.UK100', 'SPXm': '.US500', 'NDX': '.US100',
                                    'FCHI': '.F40', 'GDAXI': '.DE30',
                                    'STOXX50E': '.STOXX50', 'J225': '.JP225', 'AUS200': '.AUS200', 'HSI': '.HK50'}
        # Get the header from the table
        col_names = [th.text for th in table[0].tr.find_all('th')]

        swap_long_short = []
        for tr in table[0].find_all("tr"):
            swap_data = [float(td.text) if isfloat(td.text) else td.text for td in tr.find_all('td')]

            # print(swap_data)
            if len(swap_data) == len(col_names):
                if swap_data[0] in tradeview_to_bgi_symbols:  # Want to do the CFD conversion
                    swap_data[0] = tradeview_to_bgi_symbols[swap_data[0]]
                    # Want to do some CFD Digit Correction (This is a rough Guess, when compared to BGI Data)
                    for i in range(1, len(swap_data)):
                        if swap_data[0] in [".US500", ".DE30"]: # Looks like these 2 needs to multiply by 10
                            swap_data[i] = round(swap_data[i], 2)  # Looks like it needs to be divided by 10
                        else:
                            swap_data[i] = round(swap_data[i] * 0.1, 2)  #Looks like it needs to be divided by 10

                swap_long_short.append(dict(zip(col_names, swap_data)))
        return swap_long_short
    except:
        return {}


# Scrape if off the website.
# Not by JSON
@unsync
def get_swaps_fpmarkets():
    try:
        url = "https://www.fpmarkets.com/swap-point"
        return_data = requests.get(url=url)
        soup = BeautifulSoup(return_data.text, 'html.parser')
        table = soup.find_all("table")
        if len(table) == 0:
            return {}

        fpm_to_BGI_symbol = {'GER30': '.DE30', 'FRA40': '.F40', 'UK100': '.UK100', 'JP225': '.JP225',
                             'US30': ".US30", 'US100': '.US100',
                             'US500': '.US500', 'XTIUSD': '.USOil', 'XBRUSD': '.UKOil',
                             "HK50": ".HK50",
                             "EURO50": ".STOXX50",
                             "CHINA50": ".A50"}

        # Get the header from the table
        col_names = [th.text for th in table[0].tr.find_all('th')]

        swap_long_short = []
        for tr in table[0].find_all("tr"):
            swap_data = [float(td.text) if isfloat(td.text) else td.text for td in tr.find_all('td')]
            # print(swap_data)
            if len(swap_data) == len(col_names):
                if swap_data[0] in fpm_to_BGI_symbol:  # Doing CFD Conversion
                    swap_data[0] = fpm_to_BGI_symbol[swap_data[0]]
                swap_long_short.append(dict(zip(col_names, swap_data)))
        return swap_long_short
    except:
        return {}


# Scrape if off the website.
# Not by JSON
@unsync
def get_swaps_ebhforex():
    try:
        url = "https://ebhforex.com/faq/rollover-policy/"
        return_data = requests.get(url=url)
        soup = BeautifulSoup(return_data.text, 'html.parser')
        table = soup.find_all("table")
        if len(table) == 0:
            return {}

        ebh_to_BGI_symbol = {'D30EUR.': '.DE30', 'F40EUR.': '.F40', '100GBP.': '.UK100', '225JPY.': '.JP225',
                             'U30USD.': ".US30", 'NASUSD.': '.US100',
                             'SPXUSD.': '.US500', 'USOUSD.': '.USOil', 'UKOUSD.': '.UKOil'}

        tr = table[0].find_all('tr')  # Want to get to the 2nd row of the table.
        # Get the header from the table
        col_names_1 = [th.text for th in tr[0]]
        col_names_2 = [th.text for th in tr[1]]

        swap_long_short = []
        for tr_1 in tr[2:]:  # Want to start from the 3rd row.

            col_name_table = [td.get_attribute_list('class')[0] for td in tr_1.find_all('td')]
            long_index_list = [
                i if isinstance(col_name_table[i], str) and col_name_table[i].lower().find("long") >= 0 else -1 for i in
                range(len(col_name_table))]
            long_index = max(long_index_list)
            short_index_list = [
                i if isinstance(col_name_table[i], str) and col_name_table[i].lower().find("short") >= 0 else -1 for i
                in range(len(col_name_table))]
            short_index = max(short_index_list)

            if long_index == -1 or short_index == -1:
                continue

            # # Example of swap data.
            # ['AUDUSD.', -1.7535, -5.6805, '', '', '']
            # ['NZDUSD.', -5.124, -1.764, '', '', '']
            # ['USDCAD.', -2.163, -9.093, '', '', '']
            # ['AUDCHF.', 9.6615, -22.785, '', '', '']
            # ['EURAUD.', -65.646, 22.135, '', '', '']
            swap_data = [float(td.text) if isfloat(td.text) else td.text for td in tr_1.find_all('td')]
            # print(swap_data)

            if len(swap_data) >= 3:  # Want to do symbol conversion for CFDs
                if swap_data[0] in ebh_to_BGI_symbol:
                    swap_data[0] = ebh_to_BGI_symbol[swap_data[0]]
                else:  # Their own FX Symbol has a . we want to remove that.
                    swap_data[0] = swap_data[0].replace(".", "")

            # Need to build the symbol Dict. Need to remove . from the symbol name.
            swap_long_short.append(
                dict(zip(["Symbol", "Long", "Short"], [swap_data[0], swap_data[long_index], swap_data[short_index]])))

        return swap_long_short
    except:
        return {}


## https://www.globalprime.com/trading-conditions/swaps-financing/
##https://admin.gleneagle.com.au/Account/SwapPoints
# Not by JSON
@unsync
def get_swaps_globalprime():
    try:
        url = "https://admin.gleneagle.com.au/Account/SwapPoints"
        return_data = requests.get(url=url)
        soup = BeautifulSoup(return_data.text, 'html.parser')
        table = soup.find_all("table")
        if len(table) == 0:
            return {}

        # Symbol Translations.
        globalprime_to_BGI_Symbol = {'AUS200': '.AUS200', 'EUSTX50': '.STOXX50', 'FRA40': '.F40', 'GER30': '.DE30',
                                     'HK50': '.HK50', 'ES35' : '.ES35',
                                     'JPN225': '.JP225', 'NAS100': '.US100', 'UK100': '.UK100', 'UKOIL': '.UKOil',
                                     'US30': '.US30',  'US500': '.US500', 'XTIUSD': '.USOil'}
        swap_long_short = []
        for t in table:
            # Get the header from the table
            col_names = [th.text for th in t.tr.find_all('th')]

            # Want to dynamically find Long and Short Index
            long_index_list = [i if isinstance(col_names[i], str) and col_names[i].lower().find("long") >= 0 else -1 for
                               i in range(len(col_names))]
            long_index = max(long_index_list)
            short_index_list = [i if isinstance(col_names[i], str) and col_names[i].lower().find("short") >= 0 else -1
                                for i in range(len(col_names))]
            short_index = max(short_index_list)

            if long_index == -1 or short_index == -1:  # If cannot determine if it's a short or long index
                continue

            for tr in t.find_all("tr"):
                swap_data = [float(td.text) if isfloat(td.text) else td.text for td in tr.find_all('td')]
                # print(swap_data)
                if len(swap_data) == len(col_names):
                    digit_multiplier = 1    # If we need to do any digit corrections
                    # Want to do the symbol translation.
                    # Mainly for CFDs tho.
                    if swap_data[0] in globalprime_to_BGI_Symbol:
                        swap_data[0] = globalprime_to_BGI_Symbol[swap_data[0]]
                        # does some digit corrections.
                        if swap_data[0] in [".US500", ".DE30"]:
                            digit_multiplier = 100 # These 2 symbols seems like they need the 100 Multiplier
                        else:
                            digit_multiplier = 10  # For all other CFDs, we want to multiply the swaps by 10 to make it up to BGI Values.
                    else:
                        swap_data[0] = swap_data[0].replace(".", "")
                    swap_long_short.append(dict(zip(["Symbol", "Long", "Short"],
                                                    [swap_data[0],
                                                     round(digit_multiplier*swap_data[long_index],2),
                                                     round(digit_multiplier*swap_data[short_index],2)])))

        return swap_long_short
    except:
        return {}


# Want to use the same code as Flask.
# Will use sqlalchemy instead of py-sql
def init_SQLALCHEMY():
    # TODO: Link this with Flask's config. Else, there  might be trouble when moving SQL Bases.
    SQLALCHEMY_DATABASE_URI = 'mysql+pymysql://mt4:1qaz2wsx@192.168.64.73/aaron'
    return create_engine(SQLALCHEMY_DATABASE_URI)


# Will attempt to get from SQL (64.73). (if possible), and save it into a file.
# If not possible, will read from the file. Fail-safe should SQL not be connected as well.
def get_from_sql_or_file(sql_query_line, file_name, db):
    # """SELECT Core_Symbol, Contract_size, Digits, Type, Currency, m.Swap_markup_profile as Swap_markup_profile, Long_Markup, Short_Markup
    # FROM aaron.swap_bgicoresymbol as b, aaron.swap_markup_profile as m
    # where b.swap_markup_profile = m.Swap_markup_profile"""

    try:  # Will try to connect to SQL
        # db = init_SQLALCHEMY()
        sql_query = text(sql_query_line)
        raw_result = db.engine.execute(sql_query)  # Query the SQL
        result_data = raw_result.fetchall()  # Return Result
        result_col = raw_result.keys()  # Dict of the results
        result = pd.DataFrame(result_data, columns=result_col)  # Form into Dataframe

        # file_name = "Upload_Swaps/BGI_core_symbol.xls"
        # Want to save it, with no index.
        result.to_excel(file_name, index=False)
        return result
    except:  # If cannot, we will read the file and return the results from that excel.
        return pd.read_excel(file_name)


# Optional Parameter, backtrace_days_max=0 if we want to enforce to get today's swaps
# To enable count back of a few hour. (IF CFH uploads late, past midnight SGT)
# Since we can upload triple swaps, we do not need to divide. Enable divide if we want to compare swaps.
# Backtrace_days_max=5 # How many days to backtrace, max.
# start_date="" When do we want to start, if not today
# divide_by_days=False If we want to divide by days, so that comparing is easier.
# cfd_convertion=False to convert to BGI Symbols or not
# cfd_conversion=pd.DataFrame([]) A pandas dataframe for the digits.
@unsync
def CFH_Soap_Swaps(backtrace_days_max=5, start_date="", divide_by_days=False, cfd_conversion=False, df_cfd_conversion=pd.DataFrame([])):
    # TODO: Update Minitor Tools Table.

    wsdl_url = "https://ws.cfhclearing.com:8094/ClientUserDataAccess?wsdl"
    session_soap = Session()
    session_soap.auth = HTTPBasicAuth("BG_Michael", "Bgil8888!!")
    client = Client(wsdl_url, transport=Transport(session=session_soap))

    # Want to get the Client number that BGI has with CFH.
    client_details = client.service.GetAccounts()
    client_num = client_details[0].AccountId if len(client_details) > 0 else -1

    # Need to get the symbol decimals from CFH.
    # Since their digits and ours might be different.
    total_symbols_raw = client.service.GetInstruments()
    # Cast it to a dict.
    total_symbols = [{"InstrumentSymbol": sym["InstrumentSymbol"], "Decimals": sym["Decimals"]} for sym in total_symbols_raw if all([c in sym for c in ["InstrumentSymbol", "Decimals"]])]
    total_symbols_raw = None # Free up some space.
    df_cfh_decimal = pd.DataFrame(total_symbols)


    # Get the FX Cost.
    # Also, PM, and USOil, UKOil, Since it's considered a commodity
    cfh_fx_swaps = None
    count = 0
    # Will want to know when to start the date for swaps.
    swap_date = datetime.datetime.now().date() if start_date == "" else start_date
    while cfh_fx_swaps == None and count < backtrace_days_max:  # At max go back 5 days. Just need to get the swaps no matter what.
        # print((datetime.datetime.now() - datetime.timedelta(days=count)).date())
        swap_date = swap_date - datetime.timedelta(days=count)
        cfh_fx_swaps = client.service.GetTomNextSwapRates(client_num, swap_date)
        count = count + 1

    # Want to do the conversion of BGI/OZ Digits, CFH Decimals.
    if cfh_fx_swaps==None:  # No return. we will return an empty file.

        df_cfh_fx = pd.DataFrame([])
        fx_swaps =  df_cfh_fx.to_dict("records")
    else:   # If there are returns, meaning CFH has uploaded swaps.

        df_cfh_fx = pd.DataFrame([helpers.serialize_object(c ,dict) for c in cfh_fx_swaps]) # Convert from zeep to dict
        df_cfh_fx = df_cfh_fx.merge(df_cfh_decimal) # Will join on InstrumentSymbol

        if cfd_conversion == False:   # This is when we need OZ Symbols
            df_cfh_fx["InstrumentSymbol"] = df_cfh_fx["InstrumentSymbol"].replace({"USOUSD": "vUSOil", "UKOUSD": "vUKOil"})
            # We need to remove the "/"
            df_cfd_conversion["Symbol"] = df_cfd_conversion.apply(lambda x: x['CoreSymbol'].replace("/",""), axis = 1)
        else: # MT4 Symbols
            df_cfh_fx["InstrumentSymbol"] = df_cfh_fx["InstrumentSymbol"].replace({"USOUSD": ".USOil", "UKOUSD": ".UKOil"})

        df_cfh_fx = df_cfh_fx.merge(df_cfd_conversion, left_on="InstrumentSymbol", right_on="Symbol")

        #df_cfh_fx[df_cfh_fx["Decimals"] != df_cfh_fx["Digits"]][["Symbol", "Digits", "Decimals"]]

        # Pip is give. So multiply by 10
        # Long side needs to be multiplied by -1
        # The rule here is that if the swap rate is positive for the short side, you will receive swap and vice versa
        # If the swap rate is positive for the long side, you will pay swap and vice versa.
        # Here is another example on our product page:http://www.cfhclearing.com/products/#fx_com_calc
        df_cfh_fx["LongPosPips"] = df_cfh_fx.apply(lambda x:  round(float(-10 * x["LongPosPips"] / (10 ** (x["Decimals"] - x["Digits"]))),4), axis=1)
        df_cfh_fx["ShortPosPips"] = df_cfh_fx.apply(lambda x: round(float( 10 * x["ShortPosPips"] / (10 ** (x["Decimals"] - x["Digits"]))), 4),  axis=1)

        # We want to be mindful about the triple swaps.
        # Depends on the situation, we might need to divide, or not.
        # If it is a holiday, s['ToValueDate'] == s['FromValueDate'], so need to do Max 1
        if divide_by_days == True:  # If we want to divide by Days
            df_cfh_fx["ShortPosPips"] = df_cfh_fx.apply(lambda x: round( x["ShortPosPips"] / max((x['ToValueDate'] - x['FromValueDate']).days, 1), 4), axis=1)
            df_cfh_fx["ShortPosPips"] = df_cfh_fx.apply(lambda x: round( x["ShortPosPips"] / max((x['ToValueDate'] - x['FromValueDate']).days, 1), 4), axis=1)

        # Select Which column is needed.
        if cfd_conversion == False: # This is when we need OZ Symbols
            df_cfh_fx["Symbol Group Path"] = "*"
            df_cfh_fx=df_cfh_fx.rename(columns={"CoreSymbol" : "Core Symbol", "LongPosPips" : "Long Points", "ShortPosPips" : "Short Points"})
            df_cfh_fx = df_cfh_fx[['Core Symbol', "Symbol Group Path", 'Long Points', 'Short Points']]
        else:   # For MT4
            df_cfh_fx = df_cfh_fx[[ 'InstrumentSymbol', 'LongPosPips', 'ShortPosPips']]
            # Want to append the Symbol, Long and Short.
            df_cfh_fx = df_cfh_fx.rename(
                columns={'InstrumentSymbol': 'Symbol', 'LongPosPips': 'Long', 'ShortPosPips': 'Short'})


        fx_swaps = df_cfh_fx.to_dict("records")



    # Multiple of -1 on the long side to flip it.
    # fx_swaps = [{"Symbol": s['InstrumentSymbol'],
    #              "Long": -1 * round(float(s['LongPosPips']) if divide_by_days == False else float(s['LongPosPips']) /
    #              max( (s['ToValueDate'] - s['FromValueDate']).days, 1), 3),
    #              "Short": round(float(s['ShortPosPips']) if divide_by_days == False else float(s['ShortPosPips']) /
    #             max( (s['ToValueDate'] - s['FromValueDate']).days, 1), 3),
    #              } for s in cfh_fx_swaps if all(
    #     u in s for u in ["InstrumentSymbol", 'LongPosPips', 'ShortPosPips', 'FromValueDate', 'ToValueDate'])]



    ## ---------------------------------- CFDs ---------------------------
    cfh_cfd_swaps = None
    count = 0
    # Will want to know when to start the date for swaps.
    swap_date = datetime.datetime.now().date() if start_date == "" else start_date
    while cfh_cfd_swaps == None and count < backtrace_days_max:  # At max go back 5 days. Just need to get the swaps no matter what.
        # print((datetime.datetime.now() - datetime.timedelta(days=count)).date())
        swap_date = swap_date - datetime.timedelta(days=count)
        cfh_cfd_swaps = client.service.GetCFDCost(client_num, swap_date)
        count = count + 1

    # Want to transform the column names.
    # And to include only those that is needed.
    # Swap date in python. Monday : 0, Tuesday: 1...
    cfd_col = {"InstrumentSymbol": "Symbol", "LongPosCost": "Long", "ShortPosCost": "Short", "GrossDividend": "Dividend"}
    cfd_cost = [{bgi_col:c_swaps[col] for col, bgi_col in cfd_col.items()} for c_swaps in cfh_cfd_swaps if all([col in c_swaps for col in cfd_col])]

    # Want to get Daily Swaps
    # If it's friday, and we also want to divide...
    if divide_by_days == True and swap_date.weekday() == 4:
        for i in range(len(cfd_cost)):
            cfd_cost[i]['Long'] = float(cfd_cost[i]['Long'] / 3)
            cfd_cost[i]['Short'] = float(cfd_cost[i]['Short'] / 3)

    # Clean up, to round to 4 decimal places.
    for i in range(len(cfd_cost)):
        # If the swap rate is positive for the long side, you will pay swap and vice versa.
        # Also, to calculate in the dividend
        cfd_cost[i]['Long'] = -1 * round(float(cfd_cost[i]['Long']), 4) +  round(float(cfd_cost[i]['Dividend']), 4)     # For long, we need to invert it.
        cfd_cost[i]['Short'] = round(float(cfd_cost[i]['Short']), 4) -  round(float(cfd_cost[i]['Dividend']), 4)
        #cfd_cost[i]['Dividend'] = round(float(cfd_cost[i]['Dividend']), 4)

    # TODO: Need to understand what to do with the dividend.
    for i in range(len(cfd_cost)):
        cfd_cost[i].pop('Dividend')

    cfd_cost_return = []
    cfh_bgi_Symbols = {'200AUD': ".AUS200", 'CHN50USD': ".A50", 'F40EUR': ".F40",
                       'D30EUR': ".DE30",  'E50EUR': ".STOXX50", 'H33HKD': ".HK33",
                       '225JPY': ".JP225", 'E35EUR': ".ES35", '100GBP': ".UK100",
                       'NASUSD': ".US100",  'SPXUSD': ".US500", 'U30USD': ".US30"}



    # If we need to convert to BGI Symbols.
    if cfd_conversion == True:
        df_cfd = pd.DataFrame(cfd_cost)                                         # Put into df
        df_cfd = df_cfd[df_cfd["Symbol"].isin(cfh_bgi_Symbols)]                 # Want only those that BGI has
        df_cfd["Symbol"] = df_cfd["Symbol"].apply(lambda x: cfh_bgi_Symbols[x]  # Change Symbol to BGI Symbol
            if x in cfh_bgi_Symbols else x)


        # Return a Pandas Data Frame
        #cfd_conversion = get_MT4_cfd_Digits()
        df_cfd = df_cfd.merge(df_cfd_conversion)


        # Digits Corrections.
        df_cfd["Long"] = df_cfd.apply(lambda x:  round(x["Long"]  * 10 ** x["Digits"], 3), axis = 1)
        df_cfd["Short"] = df_cfd.apply(lambda x: round(x["Short"] * 10 ** x["Digits"], 3), axis=1)
        df_cfd = df_cfd[["Symbol", "Long", "Short"]]    # Only want specific columns
        cfd_cost_return = df_cfd.to_dict("records")


    else: # Just return those that OZ has..
        #cfd_cost = [c for c in cfd_cost if "Symbol" in c and c["Symbol"] in cfh_bgi_Symbols]
        df_cfd = pd.DataFrame(cfd_cost)
        df_cfd = df_cfd.merge(df_cfd_conversion, left_on="Symbol", right_on="CoreSymbol")



        # Does the digit corrections.
        df_cfd["Long"] = df_cfd.apply(lambda x:  round(x["Long"]  * 10 ** x["Digits"], 3), axis = 1)
        df_cfd["Short"] = df_cfd.apply(lambda x: round(x["Short"] * 10 ** x["Digits"], 3), axis=1)
        df_cfd["Symbol Group Path"] = '*'   # For OZ uploads.

        df_cfd.rename(columns={"CoreSymbol": "Core Symbol", "Long": "Long Points","Short": "Short Points"}, inplace=True)
        df_cfd = df_cfd[["Core Symbol", "Symbol Group Path" ,"Long Points", "Short Points"]]  # Only want specific columns
        cfd_cost_return = df_cfd.to_dict("records")

    return fx_swaps + cfd_cost_return


#df_cfd_conversion=get_MT4_cfd_Digits()
# For CFH Digit Calculations.
def get_MT4_cfd_Digits(db=False):
    # Need to get from SQL, the digits for MT4.
    # Because we need to multiply by the correct digits.
    if db == False:
        db = init_SQLALCHEMY()

    sql_query_line = """select Symbol, Digits
        From live1.mt4_symbols_update
        where `SECURITY` like "CFD%Oil" or 
            `SECURITY` in ("Group 5", "Exotic Pairs", "Gold", "Silver")
        order by `Symbol`"""

    # Return a Pandas Data Frame
    cfd_conversion = get_from_sql_or_file(sql_query_line, current_app.config["VANTAGE_UPLOAD_FOLDER"] + "BGI_CFD_Digits.xls", db)
    return cfd_conversion

# For the Upload to OZ.
#df_cfd_conversion=get_OZ_CFH_cfd_Digits()
# Since OZ uses digits, but CFH dosn't, we need to do some converstion.
def get_OZ_CFH_cfd_Digits():
    # Margin Class
    # Want to get margin Core symbol digits
    margin_c = OZ_Rest_Class("Margin")
    margin_symbol_setting = margin_c.get_core_symbol()

    # Want to get those that are FOREX, METAL, or CFH indicies
    cfd_conversion = [m for m in margin_symbol_setting['settings']['Symbols']
                      if "SymbolGroupPath" in m and
                      (m['SymbolGroupPath'].find("CFH") >= 0 or m['SymbolGroupPath'].find("METAL") >= 0
                       or m['SymbolGroupPath'].find("FOREX") >= 0 or m['SymbolGroupPath'].find("ENERGIES") >= 0)]

    df_cfh_indices = pd.DataFrame(cfd_conversion)
    return df_cfh_indices[[ 'CoreSymbol', 'Digits']]


# Will get other broker swaps with requests.
# and join them together.
def get_broker_swaps(db=False):

    # Using Unsync to get the details.
    # Will need to .result() to get the returned data
    # Doing it this way would cause it to be threaded.
    # Wait when it's calling .result()

    #cfh_unsync = CFH_Soap_Swaps(divide_by_days=True, df_cfd_conversion=get_MT4_cfd_Digits(db), cfd_conversion=True)


    fxdd_unsync = get_swaps_fxdd()
    forexDotCom_unsync = get_swaps_forexDotCom()
    saxo_unsync = get_swaps_saxo()
    tradeview_unsync = get_swaps_tradeview()
    globalprime_unsync = get_swaps_globalprime()
    ebhforex_unsync = get_swaps_ebhforex()
    fpmarkets_unsync = get_swaps_fpmarkets()



    df_fxdd = pd.DataFrame(fxdd_unsync.result())
    # df_fxdd['Long'] = df_fxdd['Long'].apply(lambda x: round(x, 2) if type(x) == "float" or  type(x) == "int" else x)
    # df_fxdd['Short'] = df_fxdd['Long'].apply(lambda x: round(x, 2) if type(x) == "float" or type(x) == "int" else x)
    df_fxdd = df_fxdd.rename(columns={"Long": "fxdd Long", "Short": "fxdd Short"})

    df_fdc = pd.DataFrame(forexDotCom_unsync.result())
    # df_fdc['Long'] = df_fdc['Long'].apply(lambda x: round(x, 2) if type(x) == "float" or  type(x) == "int" else x)
    # df_fdc['Short'] = df_fdc['Short'].apply(lambda x: round(x, 2) if type(x) == "float" or type(x) == "int" else x)
    df_fdc = df_fdc.rename(columns={"Long": "fdc Long", "Short": "fdc Short"})

    df_saxo = pd.DataFrame(saxo_unsync.result())
    df_saxo = df_saxo.rename(columns={"Long": "saxo Long", "Short": "saxo Short"})

    df_tradeview = pd.DataFrame(tradeview_unsync.result())
    df_tradeview = df_tradeview.rename(columns={"Long": "tv Long", "Short": "tv Short"})

    df_global_prime = pd.DataFrame(globalprime_unsync.result())
    df_global_prime = df_global_prime.rename(columns={"Long": "gp Long", "Short": "gp Short"})

    df_ebhforex = pd.DataFrame(ebhforex_unsync.result())
    df_ebhforex = df_ebhforex.rename(columns={"Long": "ebh Long", "Short": "ebh Short"})

    df_fpmarkets = pd.DataFrame(fpmarkets_unsync.result())
    df_fpmarkets = df_fpmarkets.rename(columns={"Long": "fpm Long", "Short": "fpm Short"})


    #df_cfh = pd.DataFrame(cfh_unsync.result())
    #df_cfh = df_cfh.rename(columns={"Long": "cfh Long", "Short": "cfh Short"})

    #swaps_array = [df_fxdd, df_fdc, df_saxo, df_tradeview, df_global_prime,df_ebhforex, df_fpmarkets, df_cfh]


    swaps_array = [df_fxdd, df_fdc, df_saxo, df_tradeview, df_global_prime,df_ebhforex, df_fpmarkets]


    df_return = pd.DataFrame([], columns=["Symbol"])
    how = "outer"
    for s in swaps_array:
        if "Symbol" in s and len(s) > 0:
            df_return = df_return.merge(s, on="Symbol", how=how)

    # Want to re-arrange the columns
    all_df_col = df_return.columns.tolist()
    Symbol_array = [a for a in all_df_col if a.find("Symbol") >= 0]
    Swap_long_array = [a for a in all_df_col if a.find("Long") >= 0]
    Swap_short_array = [a for a in all_df_col if a.find("Short") >= 0]
    everything_else_array = [u for u in all_df_col if u not in Symbol_array + Swap_long_array + Swap_short_array]

    # Want to re-arrange the columns.
    return df_return[Symbol_array + Swap_long_array + Swap_short_array + everything_else_array]
    # return df_return

# To calculate the swap after markup, based on the various % of the swap markups.
def swap_markup(swap_val, markup_percentage):
    x = swap_val
    if x > 0: # Positive. We are giving Client. We want to give Less.
        x = x * 0.01 * (100 - markup_percentage)
    elif x < 0: # It's negative. We want to take more from Client.
        x = x *  0.01 * (100 + markup_percentage)
    return x


def calculate_swaps_bgi(excel_data, db):

    #pd.set_option('display.max_rows', 200)

    df_bgi_excel = pd.DataFrame(excel_data)

    # Cast it to float.
    # Data came from SQL. Might be in string.
    df_bgi_excel["Long Points"] = df_bgi_excel["Long Points"].astype(float)
    df_bgi_excel["Short Points"] = df_bgi_excel["Short Points"].astype(float)

    #print(df_bgi_excel)

    # Get the Bloomberg Dividend.
    #df_dividend = pd.DataFrame(Get_Dividend())
    #print(df_dividend)

    # Get 3rd Party Swaps - UNSYNC. To save some time.
    tradeview_unsync = get_swaps_tradeview()
    globalprime_unsync = get_swaps_globalprime()

    #start_time = datetime.datetime.now()
    df_swaps_predict = predict_cfd_swaps(db, return_predict_only=True)
    #print("Time taken for prediction: {}s".format((datetime.datetime.now() - start_time).total_seconds()))
    # The dataframe column names has to be unique.
    df_swaps_predict.rename(columns={"BGI_Long" : "BGI_Predict_Long", "BGI_Short" : "BGI_Predict_Short"}, inplace=True)


    # Need to do Long point correction form the file that Vantage sent.
    df_bgi_excel["Long Points"] = df_bgi_excel["Long Points"] * -1


    #print("Session details: {}".format(current_app.config["VANTAGE_UPLOAD_FOLDER"]))

    # Get the Symbol details from SQL
    # Puts it into  the correct Folder
    df = get_from_sql_or_file("call aaron.Swap_Symbol_Details()",
                              current_app.config["VANTAGE_UPLOAD_FOLDER"] + "Swap_Symbol_Details.xlsx", db)

    # Merge the Vantage swaps with the Symbol details.
    df = df.merge(df_bgi_excel, how="left", left_on="vantage_coresymbol", right_on="Core Symbol")

    # Merge in Bloomberg Dividend.
    # df["Symbol_without_dot"] = df["bgi_coresymbol"].apply(lambda x: x.replace(".", "")) # So that we can do the merge.
    # df = df.merge(df_dividend, how="left", left_on="Symbol_without_dot", right_on="mt4_symbol")
    # df.drop( columns="Symbol_without_dot", inplace = True)  # Drop the column that we just created.
    #print(df)


    # Merge in CFD Regression Swaps Value
    df = df.merge(df_swaps_predict, how="left", left_on="bgi_coresymbol", right_on="Symbol")



    # Calculate the markup First.
    df["long_markup_value"] = df.apply(lambda x: swap_markup(x["Long Points"], x["Long_Markup"]), axis=1)
    df["short_markup_value"] = df.apply(lambda x: swap_markup(x["Short Points"], x["Short_Markup"]), axis=1)

    # Account for the Symbol Digit Correction
    df["long_markup_value_digit_correct"] = df["long_markup_value"] * (10 ** (df["BGI_digits"] - df["VANTAGE_digits"]))
    df["short_markup_value_digit_correct"] = df["short_markup_value"] * (
                10 ** (df["BGI_digits"] - df["VANTAGE_digits"]))

    # If there's no BGI digits, or Vantage Digits, this will Fire off.
    df['long_markup_value'] = np.where(df['long_markup_value_digit_correct'].isna(), df['long_markup_value'],
                                       df['long_markup_value_digit_correct'])
    df['short_markup_value'] = np.where(df['short_markup_value_digit_correct'].isna(), df['short_markup_value'],
                                        df['short_markup_value_digit_correct'])

    # Want to deal with all the fixed Values.
    df['long_markup_value_PlusFixed'] = np.where(df['BGI_fixed_long'].isna(), df['long_markup_value'], df['BGI_fixed_long'])

    df['long_markup_value_PlusFixed'] = round(df['long_markup_value_PlusFixed'], 4)

    df['short_markup_value_PlusFixed'] = np.where(df['BGI_fixed_short'].isna(), df['short_markup_value'],
                                                  df['BGI_fixed_short'])
    df['short_markup_value_PlusFixed'] = round(df['short_markup_value_PlusFixed'], 4)

    # Want to note Symbols are on fixed swaps
    df['Markup_Style'] = np.where( (~df['BGI_fixed_long'].isna()) | (~df['BGI_fixed_short'].isna()), \
                                                  "#CCD1D1", "")

    #print(df)
    #--- Want to show which

    # Want to deal with all the fixed Insti Values.
    df['long_markup_value_Plus_Insti_Fixed'] = np.where(df['BGI_fixed_insti_long'].isna(),
                                                        df['long_markup_value_PlusFixed'], df['BGI_fixed_insti_long'])
    df['long_markup_value_Plus_Insti_Fixed'] = round(df['long_markup_value_Plus_Insti_Fixed'], 4)

    df['short_markup_value_Plus_Insti_Fixed'] = np.where(df['BGI_fixed_insti_short'].isna(),
                                                         df['short_markup_value_PlusFixed'],
                                                         df['BGI_fixed_insti_short'])
    df['short_markup_value_Plus_Insti_Fixed'] = round(df['short_markup_value_Plus_Insti_Fixed'], 4)



    # Want to over-write the average with the regression CFD Swaps.
    df['avg_long'] = np.where(df['BGI_Predict_Long'].isna(), df['avg_long'],
                                        df['BGI_Predict_Long'])

    df['avg_short'] = np.where(df['BGI_Predict_Short'].isna(), df['avg_short'],
                                        df['BGI_Predict_Short'])




    custom_dict = {"FX": 0, "FX_20%": 0, "Exotic Pairs": 1, "PM": 2, "CFD": 4, "CFD_20%": 4}
    df.sort_values(by=["swap_markup_profile", "bgi_coresymbol"], key=lambda x: x.map(custom_dict), inplace=True)

    # df.sort_values(["swap_markup_profile", "bgi_coresymbol"],ascending=[False, True],  inplace=True)

    bgi_Col_Needed = ["bgi_coresymbol", "long_markup_value_PlusFixed",
                      "short_markup_value_PlusFixed", \
                      "swap_markup_profile", "long_markup_value_Plus_Insti_Fixed",
                      "short_markup_value_Plus_Insti_Fixed", "dividend", \
                      "avg_long", "avg_short", "Markup_Style"]



    df_tradeview = pd.DataFrame(tradeview_unsync.result())
    df_tradeview = df_tradeview.rename(columns={"Long": "tv Long", "Short": "tv Short"})
    #print("df_tradeview")

    #print(df_tradeview)

    df_global_prime = pd.DataFrame(globalprime_unsync.result())
    df_global_prime = df_global_prime.rename(columns={"Long": "gp Long", "Short": "gp Short"})

    df_Merge = df[bgi_Col_Needed].merge(df_tradeview, how="left", left_on="bgi_coresymbol", right_on="Symbol")
    df_Merge = df_Merge.merge(df_global_prime, how="left", left_on="bgi_coresymbol", right_on="Symbol")
    return df_Merge


# To see if there's any issue with the swap values
# To see if there's any issue with the swap values
def compare_swap_values(x, y):

    # minimum difference: White
    # different in sign : Yellow
    # significant difference : Red
    # minor difference: Orange


    color_dict = {"Red": "#E74C3C", "Yellow": "#F1C40F", \
               "Blue" : "#85C1E9", "Green": "#BCF90C", \
               "Orange": "#F39C12", "White": "#FFFFFF"}


    # We need to make sure input is number value
    if not isfloat(x) or not isfloat(y):
        return "White"

    diff = abs(x - y)  # The difference in the values.
    percent_difference_allowed_raise_warning = 50
    percent_difference_allowed_raise_info = 20
    min_allow_difference = 2  # If the swap only differs by X, it should be alright.
    min_allow_difference_opposing_sign = 0.5 # If swaps are different signs, it's okay if they are just 0.5 difference.

    # If they are different in sign, and past a certain threshold
    if ((x > 0 and y < 0) or (x < 0 and y > 0)) and diff > min_allow_difference_opposing_sign:
        return color_dict["Green"]

    if diff < min_allow_difference:  # To allow for minimum difference
        return color_dict["White"]

    # If there is a significant difference.
    if (percent_difference_allowed_raise_warning * 0.01 * min(abs(x), abs(y)) < diff) and diff > min_allow_difference:
        return color_dict["Red"]

    # If there is a minor difference.
    if (percent_difference_allowed_raise_info * 0.01 * min(abs(x), abs(y)) < diff) and diff > min_allow_difference:
        return color_dict["Orange"]

    return color_dict["White"]



# Get the day's chargable dividend from SQL
def Get_Dividend():

    increment_days = 1  #How many days to increment by. Need to go from the friday to monday.
    now = datetime.datetime.now() #- datetime.timedelta(days=2)

    if now.weekday() >= 4:   # 0=Monday, 4 = Thursday
        increment_days = 3 - (now.weekday() - 4) # To skip the weekend
    else:
        increment_days = 1 # Just the next day would be good.

    sql_query = "Select mt4_symbol, dividend ,`date` as `Date(DD-MM-YYYY)` from aaron.bloomberg_dividend WHERE Dividend != 0 AND `date` = '{}'".format( \
        (now+datetime.timedelta(days=increment_days)).strftime("%Y-%m-%d"))

    #df = get_from_sql_or_file("call aaron.Swap_Symbol_Details()", "Swap_Symbol_Details.xlsx", db)
    return query_SQL_return_record(sql_query)

    # [result_array, Column_Details] = Query_SQL(sql_query)
    #
    # to_table_data = [[".{}".format(a[0]), "{}".format(a[1]) ,a[2].strftime("%d-%m-%Y")] for a in result_array]
    # to_table_column = [a[0] for a in Column_Details]
    #
    # file_name = "BloombergDividendChargable_{}.csv".format(now.strftime("%d_%b_%Y"))  # File name with date.
    #
    # with open(file_name, "w") as f:
    #     f.write(",".join(to_table_column))
    #     f.write("\n")
    #     f.write("\n".join([",".join(a) for a in to_table_data]))




#----- To calculate and apply regression for swaps on CFDs


def get_dividend_history(db, backward_days=51):
    #     data_dividend = Query_SQL_Host("""SELECT mt4_symbol, dividend, date,  weekday(date) as `Weekday`
    #     FROM aaron.bloomberg_dividend where date > DATE_SUB(NOW(), INTERVAL {} DAY)""".format(backward_days), \
    #                                    SQL_IP, SQL_User, SQL_Password, SQL_Database)
    #     df_dividend = pd.DataFrame(data_dividend[0], columns=["mt4_symbol", "dividend", "date", "Weekday"])

    df_dividend = get_from_sql_or_file("""SELECT mt4_symbol, dividend, date,  weekday(date) as `Weekday` 
    FROM aaron.bloomberg_dividend where date > DATE_SUB(NOW(), INTERVAL {} DAY)""".format(backward_days),
                         current_app.config["VANTAGE_UPLOAD_FOLDER"] + "Dividend_History.xlsx", db)

    # Weekday 0 = Monday.

    # Backward calculate by 1 working day.
    df_dividend["Backwards_days"] = 1
    df_dividend["Backwards_days"] = np.where(df_dividend['Weekday'] == 6, 2, df_dividend["Backwards_days"])
    df_dividend["Backwards_days"] = np.where(df_dividend['Weekday'] == 0, 3, df_dividend["Backwards_days"])

    # Cast to date
    df_dividend["date"] = pd.to_datetime(df_dividend["date"])
    # Add the . for the symbols
    df_dividend["mt4_symbol"] = df_dividend["mt4_symbol"].apply(lambda x: ".{}".format(x))

    df_dividend["Date_merge"] = df_dividend.apply(lambda x: x["date"] - datetime.timedelta(days=x["Backwards_days"]),
                                                  axis=1)

    df_dividend_ret = df_dividend.groupby(["mt4_symbol", "Date_merge"]).sum().reset_index()

    # Cast the dividend to float
    df_dividend_ret["dividend"] = df_dividend_ret["dividend"].astype(float)

    df_dividend_ret = df_dividend_ret[["mt4_symbol", "Date_merge", "dividend"]]
    return df_dividend_ret

# To calculate what should be the swaps, with dividend incoperated into it. Both long and short.
def calculate_CFD_long_short_dividend(df):
    # Get all the dividend data that are history
    df_regression_data = df[(~df["BGI_Long"].isna()) & (~df["dividend"].isna()) & (~df["BGI_Short"].isna())]
    if len(df_regression_data) == 0:
        print("Data is Empty or NaN")
        return (0, 0)

    # Want to take the dividend for non-fridays and fridays differently.
    dividend_today = df[df["Date_merge"] == datetime.datetime.now().strftime("%Y-%m-%d")][
        ["Dividend_Friday", "Dividend_Not_Friday"]]

    #print(dividend_today.values)

    # Taking two "types" of regression for multiple linear regression.
    df_s = df_regression_data[["Dividend_Friday", "Dividend_Not_Friday", "BGI_Long", "BGI_Short"]]

    X = df_s.iloc[:, :-2].values  # Get the Dividend
    Y_Long = df_s.iloc[:, 2].values  # Get the LONG for BGI.
    Y_Short = df_s.iloc[:, 3].values  # Get the SHORT for BGI.
    # print(Y_Long)

    return (linear_regression(X, Y_Long, dividend_today.values), linear_regression(X, Y_Short, dividend_today.values))

# To run the Linear Regression for the CFD Swaps.
def linear_regression(X, Y, Predict):
    lr = LinearRegression()
    model_Long = lr.fit(X, Y)
    predictions = lr.predict(Predict)
    return predictions[0]


def get_swap_history(db, backward_days=50):
    # data = Query_SQL_Host("""SELECT Core_Symbol as `Symbol`, BGI_Long, BGI_Short, Date
    #     FROM test.bgi_swaps WHERE CORE_SYMBOL LIKE '.%' AND DATE > DATE_SUB(NOW(),
    # INTERVAL {} DAY) ORDER BY date""".format(backward_days), SQL_IP, SQL_User, SQL_Password, SQL_Database)

    SQL_Query = """SELECT Core_Symbol as `Symbol`, BGI_Long, BGI_Short, Date 
    FROM test.bgi_swaps WHERE CORE_SYMBOL LIKE '.%' AND 
    DATE > DATE_SUB(NOW(),INTERVAL {} DAY) ORDER BY date""".format(backward_days).replace("\n", " ")

    #print("\n\n{}\n\n".format(SQL_Query))

    df = get_from_sql_or_file(SQL_Query, current_app.config["VANTAGE_UPLOAD_FOLDER"] + "Swap_Value_History.xlsx", db)

    #df = pd.DataFrame(query_SQL_return_record(text(SQL_Query)))

    #df = pd.DataFrame(data[0], columns=["Symbol", "BGI_Long", "BGI_Short", "Date"])
    df["BGI_Long"] = df["BGI_Long"].astype(float)
    df["BGI_Short"] = df["BGI_Short"].astype(float)

    df["Date"] = pd.to_datetime(df["Date"])
    return df


def merge_dividend_swaps(df, df_dividend):
    # Merge and try to get a column with all the data.
    df = df.merge(df_dividend, left_on=["Symbol", "Date"], right_on=["mt4_symbol", "Date_merge"], how="outer")
    df["Symbol"] = np.where(df["mt4_symbol"].isna(), df["Symbol"], df["mt4_symbol"])
    df["Date"] = np.where(df["Date"].isna(), df["Date_merge"], df["Date"])

    #df["Weekday"] = pd.to_datetime(df["Date_merge"]).dt.weekday

    return df


# Has the ability to return only the predicted df.
def predict_cfd_swaps(db, return_predict_only=True):
    df_dividend = get_dividend_history(db, 101)
    df_swapHistory = get_swap_history(db, 100)
    df = merge_dividend_swaps(df_swapHistory, df_dividend)
    df.sort_values("Date", inplace=True)

    # Create Feature. Since Fridays usually has 3 time swaps.
    df["Date_weekday"] = df["Date_merge"].dt.weekday  # 0 = Monday, 4 = Friday

    # Want only dates that are monday (0) - Friday (4)
    df = df[df["Date_weekday"].isin([0,1,2,3,4])]

    df["Dividend_Friday"] = np.where(df["Date_weekday"] == 4, df["dividend"], 0)
    df["Dividend_Not_Friday"] = np.where(df["Date_weekday"] != 4, df["dividend"], 0)


    # Want those that are not empty
    all_symbol = df[~df["dividend"].isna()]["Symbol"].unique().tolist()

    # Want those that are only within monday - Fridays.

    for s in all_symbol:
        predict_long, predict_short = calculate_CFD_long_short_dividend(df[df["Symbol"] == s])
        # print("{} {} {}".format(s, predict_long, predict_short))
        df.loc[(df.Symbol == s) & (
                    df["Date_merge"] == datetime.datetime.now().strftime("%Y-%m-%d")), 'BGI_Long'] = predict_long
        df.loc[(df.Symbol == s) & (
                    df["Date_merge"] == datetime.datetime.now().strftime("%Y-%m-%d")), 'BGI_Short'] = predict_short

    # Need to round off the values to 2
    df['BGI_Long'] = round(df['BGI_Long'], 4)
    df['BGI_Short'] = round(df['BGI_Short'], 4)


    if return_predict_only:
        return df[df["Date_merge"] == datetime.datetime.now().strftime("%Y-%m-%d")]

    return df




def process_validated_swaps(all_data):

    # Cast it into a df
    # Hidden data is so that we can check what has been changed.
    df = pd.DataFrame(all_data, columns=["Core Symbol (BGI)", "Long Points (BGI)", "Short Points (BGI)",
                                         "Insti Long Points (BGI)", "Insti Short Points (BGI)",
                                         "Long_Hidden", "Short_Hidden" ])

    # Trying to cast to float so that it will be saved as float.. hopefully?
    for c in ["Long Points (BGI)", "Short Points (BGI)", "Insti Long Points (BGI)",
              "Insti Short Points (BGI)", "Long_Hidden", "Short_Hidden"]:
        if c in df:
            df[c] = df[c].astype(float)

    df.sort_values("Core Symbol (BGI)", inplace=True)

    # First, we want to check if anything has changed manually.
    # # If there has been, we will want to know.
    df_changed = df[(df["Long Points (BGI)"] != df["Long_Hidden"]) | (df["Short Points (BGI)"] != df["Short_Hidden"])]
    print(df_changed)

    for s in df_changed["Core Symbol (BGI)"].unique():
        flash("{} swap details were changed manually.".format(s))


    # If there has been swaps that were manually changed, we would wanna change it on the Insi as well as the Retail side.

    if len(df_changed) > 0: # There has been some changes

        # Need to ensure that the Insti Fixed Isn't changed
        # If it's not the ones that are fixed (For insti), it will follow the long/short points that were changed manually.
        df["Insti Long Points (BGI)"] = np.where(df["Insti Long Points (BGI)"] != df["Long_Hidden"],
                                                 df["Insti Long Points (BGI)"], df["Long Points (BGI)"])

        df["Insti Short Points (BGI)"] = np.where(df["Insti Short Points (BGI)"] != df["Short_Hidden"],
                                                 df["Insti Short Points (BGI)"], df["Short Points (BGI)"])

        #
        # pd.set_option('display.max_rows', 500)
        # print(df)
        #
        #
        # ## Trying to use Pandas to write to excel.
        # df_retail = df[["Core Symbol (BGI)", "Long Points (BGI)", "Short Points (BGI)"]]
        # df_insti = df[["Core Symbol (BGI)", "Insti Long Points (BGI)", "Insti Short Points (BGI)"]].rename(
        #     columns={"Insti Long Points (BGI)": "Long Points (BGI)",
        #              "Insti Short Points (BGI)": "Short Points (BGI)"})
        #
        # with pd.ExcelWriter(current_app.config["SWAPS_MT4_UPLOAD_FOLDER"] + \
        #                         'MT4Swaps {dt.day} {dt:%b} {dt.year}.xls'.format(
        #         dt=datetime.datetime.now())) as writer:
        #     df_retail.to_excel(writer, sheet_name='retail', index=False)
        #     df_insti.to_excel(writer, sheet_name='insti', index=False)



    df_changed = None  # Clear.

    #return


    # We want to upload to MT4 First.

    # Get the data into a Bracket format to be ready to inset into SQL
    # We will do a 2 hours back so that swaps can still be uploaded near midnight.
    swap_insert_list = ["(" + ",".join(["'{}'".format(x) for x in X]) + ", DATE(DATE_SUB(now(), INTERVAL 2 HOUR)))" \
                        for X in df[["Core Symbol (BGI)", "Long Points (BGI)", "Short Points (BGI)"]].values.tolist()]
    swap_insert_str = " , ".join(swap_insert_list)

    # -------------------------------Insert into Risk 64.73 Database (Aaron Database)

    sql_header_risk = "INSERT INTO aaron.bgi_Swaps ( Core_Symbol, bgi_long, bgi_short, Date ) Values  "
    footer = " ON DUPLICATE KEY UPDATE bgi_long = Values(bgi_long), bgi_short = Values(bgi_short)"

    Insert_into_sql("{} {} {}".format(sql_header_risk, swap_insert_str, footer))  # Go insert into SQL.

    flash("Risk (64.73) Swaps Insert Successful. [aaron.bgi_Swaps]")

    # -------------------------------Insert into Risk 64.73 Database (Risk Test DataBase)
    sql_header_test = "INSERT INTO test.bgi_Swaps ( Core_Symbol, bgi_long, bgi_short, Date ) Values  "
    footer = " ON DUPLICATE KEY UPDATE bgi_long = Values(bgi_long), bgi_short = Values(bgi_short)"

    risk_sql_upload_unsync = async_sql_insert(app=current_app._get_current_object(), header=sql_header_test,
                                              values=[swap_insert_str], footer=footer, sql_max_insert=500)

    # ----------------------------------------Insert into BO DB
    sql_query_bo = "INSERT INTO bgiswap.table_swap ( bgi_symbol, bgi_long, bgi_short, Update_Date ) Values  " + swap_insert_str
    # #To make it to SQL friendly text.
    raw_insert_result = db.session.execute(text("delete from bgiswap.table_swap"),
                                           bind=db.get_engine(current_app, 'bo_swaps'))
    raw_insert_result = db.session.execute(text(sql_query_bo), bind=db.get_engine(current_app, 'bo_swaps'))
    db.session.commit()  # Since we are using session, we need to commit.
    flash("BO Swaps Insert Successful.")



    flash("Swaps uploading to MT4/5. An Email will be sent when it's done.")
    upload_swaps_mt_servers(df, current_app.config["SWAPS_MT4_UPLOAD_FOLDER"], \
                            current_app.config["SWAPS_MT5_LIVE1_UPLOAD_FOLDER"], \
                            current_app.config["SWAPS_MT5_LIVE2_UPLOAD_FOLDER"], current_user.id, current_user.email)

    return


@async_fun
def upload_swaps_mt_servers(df, mt4_base_folder, mt5_L1_base_folder, mt5_L2_base_folder, username, uploader_email ):


    #retail_sheet = [["Core Symbol (BGI)",	"Long Points (BGI)", "Short Points (BGI)"]] + df[["Core Symbol (BGI)", "Long Points (BGI)", "Short Points (BGI)"]].values.tolist()
    #insti_sheet = [["Core Symbol (BGI)",	"Long Points (BGI)", "Short Points (BGI)"]] + df[["Core Symbol (BGI)", "Insti Long Points (BGI)", "Insti Short Points (BGI)"]].values.tolist()

    c_run_results = []
    email_result_dict = {}


    # ----------------------------------------- Need to upload to MT4

    # content = {'retail': retail_sheet, 'insti': insti_sheet }
    #
    # # Save the file as an Excel first.
    # # pip install pyexcel-xls
    # pe.save_book_as(bookdict = content,
    # dest_file_name = mt4_base_folder + 'MT4Swaps {dt.day} {dt:%b} {dt.year}.xls'.format(dt=datetime.datetime.now()))

    ## Trying to use Pandas to write to excel.
    df_retail = df[["Core Symbol (BGI)", "Long Points (BGI)", "Short Points (BGI)"]]
    df_insti = df[["Core Symbol (BGI)", "Insti Long Points (BGI)", "Insti Short Points (BGI)"]].rename(
                                            columns={"Insti Long Points (BGI)" : "Long Points (BGI)",
                                                "Insti Short Points (BGI)": "Short Points (BGI)"})

    with pd.ExcelWriter(mt4_base_folder + 'MT4Swaps {dt.day} {dt:%b} {dt.year}.xls'.format(dt=datetime.datetime.now())) as writer:
        df_retail.to_excel(writer, sheet_name='retail', index=False)
        df_insti.to_excel(writer, sheet_name='insti', index=False)


    # Trying to save the data into an excel using openpyxl
    #
    # workbook = Workbook()
    # sheet = workbook.active
    # sheet.title = "retail"        # Rename the sheet
    #
    # # First, we add the data into the Excel.
    # for row in dataframe_to_rows(df[["Core Symbol (BGI)", "Long Points (BGI)", "Short Points (BGI)"]],
    #                              index=False, header=True):
    #     sheet.append(row)
    #
    # df_insti = df[["Core Symbol (BGI)", "Insti Long Points (BGI)", "Insti Short Points (BGI)"]].rename(columns={"Insti Long Points (BGI)" : "Long Points (BGI)",
    #                                                "Insti Short Points (BGI)": "Short Points (BGI)"})
    #
    # insti_sheet = workbook.create_sheet("insti")  # Insti Sheet
    # # First, we add the data into the Excel.
    # for row in dataframe_to_rows(df_insti, index=False, header=True):
    #     insti_sheet.append(row)
    # workbook.save(filename= mt4_base_folder + 'MT4Swaps {dt.day} {dt:%b} {dt.year}_1.xls'.format(dt=datetime.datetime.now()))


    # Run the C++ Prog for the Upload.

    # Run the C prog to upload Swaps to MT5 Live 1
    MT4_run_res, email_result_dict["MT4_Upload"] =  run_meta_swap_upload(prog_name="Swaps_Upload_NoWait.exe",
                                                                                  cwd=mt4_base_folder,
                                                                                  server_name="MT4 Live/Demo",
                                                            c_default_return=1) # The MT4 C prog default return is 1.





    c_run_results.append(MT4_run_res)



    # Need to tidy up the excel files into the archive folder
    clean_up_folder(mt4_base_folder, file_header="mt4swaps")


    # ------------------------------ Need to upload to MT5 - Live 1---------------------------
    #content_mt5_live1 = {'ALL': retail_sheet}

    # Save the file as an Excel first.
    df_retail.to_excel(mt5_L1_base_folder + 'MT5 Swaps {dt.day} {dt:%b} {dt.year}.xls'.format(dt=datetime.datetime.now()),
                       sheet_name='ALL', index=False)


    # pip install pyexcel-xls
    # pe.save_book_as(bookdict = content_mt5_live1,
    # dest_file_name = mt5_L1_base_folder + 'MT5 Swaps {dt.day} {dt:%b} {dt.year}.xls'.format(dt=datetime.datetime.now()))
    #
    #

    # Run the C prog to upload Swaps to MT5 Live 1
    MT5_L1_run_res, email_result_dict["MT5_Live1_Upload"] =  run_meta_swap_upload(prog_name="Upload_Swaps_MT5.exe",
                                                                                  cwd=mt5_L1_base_folder,
                                                                                  server_name="MT5 Live 1")
    c_run_results.append(MT5_L1_run_res)

    # Run the C prog to upload Swaps to MT5 Demo 1
    MT5_D1_run_res, email_result_dict["MT5_Demo1_Upload"] =  run_meta_swap_upload(prog_name="Upload_Swaps_MT5_DEMO.exe",
                                                                                  cwd=mt5_L1_base_folder,
                                                                                  server_name="MT5 Demo 1")
    c_run_results.append(MT5_D1_run_res)

    # Need to tidy up the excel files into the archive folder
    clean_up_folder(mt5_L1_base_folder, file_header="mt5 swaps")


    # ------------------------------ Need to upload to MT5 - Live 2 (UK)---------------------------



    # Save the file as an Excel first.
    # pip install pyexcel-xls
    # pe.save_book_as(bookdict = content_mt5_live2,
    # dest_file_name = mt5_L2_base_folder + 'MT5 Swaps {dt.day} {dt:%b} {dt.year}.xls'.format(dt=datetime.datetime.now()))
    #

    # Save the file as an Excel first.
    df_retail.to_excel(mt5_L2_base_folder + 'MT5 Swaps {dt.day} {dt:%b} {dt.year}.xls'.format(dt=datetime.datetime.now()),
                       sheet_name='ALL', index=False)

    # Run the C prog to upload Swaps to MT5 Live 2
    MT5_L2_run_res, email_result_dict["MT5_Live2_Upload"] =  run_meta_swap_upload(prog_name="Upload_Swaps_MT5_UK.exe",
                                                                                  cwd=mt5_L2_base_folder,
                                                                                  server_name="MT5 Live 2 [UK]")
    c_run_results.append(MT5_L2_run_res)

    # Run the C prog to upload Swaps to MT5 Demo 2
    MT5_D2_run_res, email_result_dict["MT5_Demo2_Upload"] =  run_meta_swap_upload(prog_name="Upload_Swaps_MT5_UK_Demo.exe",
                                                                                  cwd=mt5_L2_base_folder,
                                                                                  server_name="MT5 Demo 2 [UK]")
    c_run_results.append(MT5_D2_run_res)


    # C_Return_Val_mt5_2, output_mt5_2, err_mt5_2  = Run_C_Prog(Path="Upload_Swaps_MT5_UK.exe", cwd=mt5_L2_base_folder)
    #
    #
    # if C_Return_Val_mt5_2 == 0:
    #     c_run_results.append(["MT5 Live 2", "Swaps uploaded Successfully.", C_Return_Val_mt5_2])
    # else:
    #     c_run_results.append(["MT5 Live 2", "Swaps upload Error: {}.".format(err_mt5_2), C_Return_Val_mt5_2])
    #
    # email_result_dict["MT5_Live2_Upload"] =  create_email_virtual_file(output_mt5_2.decode("utf-8"))
    #
    # C_Return_Val_mt5_2D, output_mt5_2D, err_mt5_2D  = Run_C_Prog(Path="Upload_Swaps_MT5_UK_Demo.exe", cwd=mt5_L2_base_folder)
    #
    # if C_Return_Val_mt5_2D == 0:
    #     c_run_results.append(["MT5 Demo 2 [UK]", "Swaps uploaded Successfully.", C_Return_Val_mt5_2D])
    # else:
    #     c_run_results.append(["MT5 Demo 2 [UK]", "Swaps upload Error: {}.".format(err_mt5_2D), C_Return_Val_mt5_2D])

    # Create the virtual file to be uploaded
    # email_result_dict["MT5_Demo2_Upload"] =  create_email_virtual_file(output_mt5_2D.decode("utf-8"))

    # Need to tidy up the excel files into the archive folder
    clean_up_folder(mt5_L2_base_folder, file_header="mt5 swaps")


    # Send email to say that upload is done.
    Send_Email(To_recipients=[uploader_email], cc_recipients=["aaron.lim@blackwellglobal.com"],
               Subject="Swaps Upload [{}]".format(datetime.datetime.now().strftime("%Y-%m-%d")),
               HTML_Text="""{Email_Header}Hi {username},<br><br>Swaps upload results for today:  
                                    {table}<br><br>Kindly find the upload logs attached.
                                       <br><br>This Email was generated at: {datetime_now} (SGT)<br><br>Thanks,<br>Aaron{Email_Footer}""".format(
                           username=username,
                           Email_Header=Email_Header,
                           table=Array_To_HTML_Table(Table_Header=["Server", "Results", "Return Code"], Table_Data=c_run_results),
                           datetime_now=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                           Email_Footer=Email_Footer),
            Attachment_Name=[],
            virtual_file=email_result_dict)

    return

# Will clean up the folders by clearing all the excel file into the "year" folder
# base_folder - The folder to clean up
# file_header - Start of the file name.
def clean_up_folder(base_folder, file_header):
    # For moving Files around.
    folder_files = [f for f in listdir(base_folder) if isfile(join(base_folder, f))]
    files_to_move = [f for f in folder_files if f.lower().find(file_header) >= 0]

    # Need to ensure that there is an Archive folder. Else, Create it.
    # If there isn't a folder with the year on it.
    file_archive_folder = base_folder + "{}".format(datetime.datetime.now().strftime("%Y"))
    #print("Creating folder: " + file_archive_folder)
    if not os.path.isdir(file_archive_folder):
        os.mkdir(file_archive_folder)         #Create the folder.

    for f in files_to_move:
        file_new_path = "{}\{}".format( file_archive_folder, f)
        if os.path.isfile(file_new_path):        #Check if there's already that file in the archive folder
            os.remove(file_new_path)    # If there is, delete it.
        # Move the new file in.
        os.rename("{}\{}".format(base_folder, f), "{}\{}".format( file_archive_folder, f))

    return

# To run and upload the swaps to the servers.
# prog_name = "Upload_Swaps_MT5.exe"
# cwd = mt5_L1_base_folder
def run_meta_swap_upload(prog_name, cwd, server_name, c_default_return=0):

    print("Running C for {}".format(server_name))
    C_Return_Val, output, err  = Run_C_Prog(Path=prog_name, cwd=cwd)

    if C_Return_Val == c_default_return:
        c_run_results =[ server_name, "Swaps uploaded Successfully.", C_Return_Val]
    else:
        c_run_results =[ server_name, "Swaps upload Error: {}.".format(err), C_Return_Val]

    # Create the virtual file to be uploaded
    f = create_email_virtual_file(output.decode("utf-8"))
    print("Done... Running C for {}".format(server_name))

    return [c_run_results, f]



# Need to generate the file with some formatting and styling.
# Using openpyxl to do so.
def generate_pretty_Swap_file(df, sheet, sheet_name, header_style):


    thin = Side(border_style="thin", color="000000")

    sheet.title = sheet_name        # Rename the sheet

    # First, we add the data into the Excel.
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)

    # Now let's apply this to all first row (header) cells
    header_row = sheet[1]


    for cell in header_row:
        cell.style = header_style

    for row in sheet:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # We know that this would be 3. But we'll check it anyway
    for i in range(len(sheet[1])):
        sheet.column_dimensions[get_column_letter(i + 1)].width = 18

    return sheet
