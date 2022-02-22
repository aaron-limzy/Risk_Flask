from flask import Blueprint, render_template, Markup, url_for, request, current_app, jsonify, flash, redirect, Response
from flask_login import current_user, login_user, logout_user, login_required

from Aaron_Lib import *
from Helper_Flask_Lib import *
from app.Swaps.A_Utils import *
from sqlalchemy import text
from app.extensions import db, excel

from app.Swaps.forms import *

from app.Swaps.get_swaps_all import *

from flask_table import create_table, Col

import requests

import pyexcel
from werkzeug.utils import secure_filename
from app.decorators import roles_required
from app.background import *

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from flask import make_response


from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter

from app.Swaps.table import *
#from formencode import variabledecode


swaps = Blueprint('swaps', __name__)

# @swaps.before_request
# def log_request_info():
#     current_app.logger.info('{} - Trying logging for all swaps Blueprint.'.format(request.remote_addr))
#     print(request)
#     current_app.logger.info('Headers: %s', request.headers)
#     current_app.logger.info('Body: %s', request.get_data())


    #print("Before each request...")


# Want to log the use of the page.
@swaps.before_request
def before_request():

    # Don't want to record any ajax calls.
    endpoint = "{}".format(request.endpoint)
    if endpoint.lower().find("ajax") >=0:
        return
    else:

        # check if the user is logged.
        if not current_user.is_authenticated:
            return
        raw_sql = "INSERT INTO aaron.Aaron_Page_History (login, IP, full_path, datetime) VALUES ('{login}', '{IP}', '{full_path}', now()) ON DUPLICATE KEY UPDATE datetime=now()"
        sql_statement = raw_sql.format(login=current_user.id,
                                       IP=request.remote_addr,
                                       full_path=request.full_path)

        async_sql_insert_raw(app=current_app._get_current_object(),
                             sql_insert=sql_statement)



@swaps.route('/Swaps/BGI_Swaps')
@roles_required(["Risk", "Admin", "Risk_TW", "Risk_UK"])
def BGI_Swaps():
    description = Markup("Swap values uploaded onto MT4/MT5. <br>\
   Swaps would be charged on the roll over to the next day.<br> \
    Three day swaps would be charged for FX on weds and CFDs on fri.<br>" +
                         "Swaps are saved on 64.73 test.bgi_swaps table")

    return render_template("Standard_Single_Table.html", backgroud_Filename='css/Faded_car.jpg', Table_name="BGI_Swaps", \
                           title="BGISwaps", ajax_url=url_for("swaps.BGI_Swaps_ajax"),
                           description=description, replace_words=Markup(["Today"]))



@swaps.route('/Swaps/BGI_Swaps_ajax', methods=['GET', 'POST'])
@roles_required(["Risk", "Admin", "Risk_TW", "Risk_UK"])
def BGI_Swaps_ajax():     # Return the Bloomberg dividend table in Json.

    start_date = get_working_day_date(datetime.date.today(), -5)
    sql_query = text("SELECT * FROM aaron.`bgi_swaps` where date >= '{}' ORDER BY Core_Symbol, Date".format(start_date.strftime("%Y-%m-%d")))
    raw_result = db.engine.execute(sql_query)
    result_data = raw_result.fetchall()
    result_col = raw_result.keys()
    return_result = [dict(zip(result_col, d)) for d in result_data]



    # Want to get unuqie dates, and sort them.
    unique_date = list(set([d['Date'] for d in return_result if 'Date' in d]))
    unique_date.sort()

    # Want to get unuqie Symbol, and sort them.
    unique_symbol = list(set([d['Core_Symbol'] for d in return_result if 'Core_Symbol' in d]))
    unique_symbol.sort()


    swap_total = [] # To store all the symbol
    for s in unique_symbol:
        swap_long_buffer = dict()
        swap_short_buffer = dict()
        swap_long_buffer['Symbol'] = s  # Save symbol name
        swap_long_buffer['Direction'] = "Long"
        swap_short_buffer['Symbol'] = s
        swap_short_buffer['Direction'] = "Short"
        for d in unique_date:       # Want to get Values of short and long.
            swap_long_buffer[d.strftime("%Y-%m-%d (%a)")] = find_swaps(return_result, s, d, "bgi_long")
            swap_short_buffer[d.strftime("%Y-%m-%d (%a)")] = find_swaps(return_result, s, d, "bgi_short")
        swap_total.append(swap_long_buffer)
        swap_total.append(swap_short_buffer)

    return json.dumps(swap_total)


@swaps.route('/Bloomberg_Dividend')
@roles_required(["Risk", "Risk_TW","Risk_UK", "Admin"])
def Bloomberg_Dividend():
    description = Markup("Dividend Values in the table above are 1-day early, when the values are uploaded as swaps onto MT4. <br>\
    Dividend would be given out/charged the next working day.")
    return render_template("Standard_Single_Table.html", backgroud_Filename='css/Charts.jpg', Table_name="CFD Dividend", \
                           title="CFD Dividend", ajax_url=url_for("swaps.Bloomberg_Dividend_ajax"),
                           description=description, replace_words=Markup(["Today"]))



@swaps.route('/Bloomberg_Dividend_ajax', methods=['GET', 'POST'])
@roles_required(["Risk", "Risk_TW","Risk_UK", "Admin"])
def Bloomberg_Dividend_ajax():     # Return the Bloomberg dividend table in Json.

    start_date = get_working_day_date(datetime.date.today(), -3)
    end_date = get_working_day_date(datetime.date.today(), 5)

    sql_query = text("Select * from aaron.bloomberg_dividend WHERE `date` >= '{}' and `date` <= '{}' AND WEEKDAY(`date`) BETWEEN 0 AND 4 ORDER BY date".format(start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")))

    raw_result = db.engine.execute(sql_query)
    result_data = raw_result.fetchall()     # Return [('A50', 'XIN9I Index', 0.0, datetime.date(2019, 8, 14), datetime.datetime(2019, 8, 14, 8, 0, 14)), ....]

    # Want to get all the weekdays in the middle of start_date and end_date.
    all_dates = [start_date + datetime.timedelta(days=d) for d in range(abs((end_date - start_date).days)) \
                 if (start_date + datetime.timedelta(days=d)).weekday() in range(5)]

    # dict of the results
    result_col = raw_result.keys()
    result_dict = [dict(zip(result_col,d)) for d in result_data]

    # Get unique symbols using set, then sort them.
    symbols = list(set([rd[0] for rd in result_data]))
    symbols.sort()
    return_data = []   # The data to be returned.

    for s in symbols:
        symbol_dividend_date = dict()
        symbol_dividend_date["Symbol"] = s
        for d in all_dates:
            symbol_dividend_date[get_working_day_date(start_date=d, weekdays_count=-1).strftime( \
                "%Y-%m-%d")] =  find_dividend(result_dict, s, d)
        return_data.append(symbol_dividend_date)
    return json.dumps(return_data)



@swaps.route('/Swaps/upload_LP_Swaps', methods=['GET', 'POST'])
@roles_required(["Risk", "Risk_TW","Risk_KH", "Admin", "Dealing"])
def upload_Swaps_csv():


    title = "Swaps Upload"
    header = "Swaps Upload"
    description = Markup("""Upload Swap files from LP [Vantage].<br>Kindly ensure that the file is correct and includes the following columns.<br>
                        - Core Symbol<br>
                        - Long Points<br>
                        - Short Points<br>""")
    form = UploadForm()

    session_to_pop = ["swap_validated_data_datetime", "swap_validated_data", "Swap_excel_upload"]

    for s in session_to_pop:
        if s in session:
            session.pop(s, None)


    if request.method == 'POST' and form.validate_on_submit():

        # The columns to check.
        col_to_check = ['Core Symbol', 'Symbol Group Path', 'Long Points', 'Short Points']

        # To track which columns are missing.
        missing_col = []

        record_dict = request.get_records(field_name='upload', name_columns_by_row=0)
        #print(record_dict)

        # We want to check the Valtage file to access if it's usable.
        #pd.set_option('display.max_rows', None)
        # Get the dataframe of the records.
        df = pd.DataFrame(record_dict)
        #print(df)
        # Want to check how many "long" and "Short" Columns there are.
        error_found = 0    # Counter for any issues on the file.

        # Checking for missing, or double/multiple entries of the same columns
        for x in ["long", "short", "core symbol"]:
            if sum([c.lower().find(x) == 0 for c in df.columns]) > 1:
                error_found = error_found + 1
                flash(Markup("More than 1 <b>'{}'</b> column found.".format(x)))
            elif sum([c.lower().find(x) == 0 for c in df.columns]) < 1:
                error_found = error_found + 1
                flash(Markup("No <b>'{}'</b> column found.".format(x)))

        # Check for main columns to see if it's missing.
        if not all([c in df.columns for c in col_to_check]):
            error_found = error_found + 1
            missing_col = ["<b>{}</b>".format(c) for c in col_to_check if c not in df]
            flash(Markup("Columns missing from excel: {}.".format(" & ".join(missing_col))))

        # Check if the Symbol is empty.
        # Because it's taken from Flask, the NAN turns to ""
        if error_found == 0: # If there are already errors, we will not be able to check that if the columns are not around
            df_na_symbol = df[(df['Core Symbol'].isna()) | (df['Core Symbol'] == "")]
            #print("df_na_symbol: ")
            #print(df_na_symbol)
            na_symbol_index = [c + 2 for c in df_na_symbol.index.to_list()]
            for n in na_symbol_index:
                error_found = error_found + 1 # Increment Error Count
                flash(Markup("Row {} in csv file is missing Core Symbol.".format(n)))

        # Check if there are blanks in the LONG/SHORT Columns
        # Because it's taken from Flask, the NAN turns to ""
            df_na = df[(df['Long Points'].isna()) | (df['Long Points']=="") |
                        (df['Short Points'].isna()) | (df['Short Points']=="")] # Check for Blanks


            for s in df_na["Core Symbol"].to_list():
                error_found = error_found + 1 # Increment Error Count
                flash(Markup("<b>{}</b> in csv file is <u>missing Long/Short Points</u>".format(s)))

            # Check if there the LONG/SHORT Columns are all just numbers
            df_not_float = df[ df["Long Points"].apply(lambda x: not isfloat(str(x))) |
                        df["Short Points"].apply(lambda x: not isfloat(str(x)))] # Check for non-float
            for s in df_not_float["Core Symbol"].to_list():
                error_found = error_found + 1 # Increment Error Count
                flash(Markup("<b>{}</b> : Long/Short points <u>isn't a number.</u>".format(s)))

        if error_found == 0 :   ## If Only there are no issues.
            start_time = datetime.datetime.now()
            df["Core Symbol"] = df["Core Symbol"].apply(lambda x: "'{}'".format(x))  # Cast to string. Add the '
            df["Long Points"] = df["Long Points"].apply(lambda x: "'{}'".format(x))
            df["Short Points"] = df["Short Points"].apply(lambda x: "'{}'".format(x))
            df["Date"] = "'{}'".format(datetime.datetime.now().strftime("%Y-%m-%d"))  # Write in system time.

            # Getting the data to write to SQL for Vantage Raw.
            data = df[['Core Symbol', 'Long Points', 'Short Points', 'Date']].values.tolist()
            sql_data = " , ".join(["({})".format(",".join(d)) for d in data])

            # SQL Statement preparation for inset.
            sql_header = "INSERT INTO aaron.swaps_vantage_raw (`core_symbol`, `vantage_long`, `vantage_short`, `date`) VALUES "
            sql_footer = " ON DUPLICATE  KEY UPDATE `vantage_long`=VALUES(`vantage_long`), `vantage_short`=VALUES(`vantage_short`)  "

            # To construct the sql statement. header + values + footer.
            sql_insert = sql_header + sql_data + sql_footer
            Insert_into_sql(sql_insert) # Go insert into SQL.
            print("Time Taken to send swaps to SQL: {}".format((datetime.datetime.now() - start_time).total_seconds()))
            return redirect(url_for('swaps.Swap_upload_form'))


    return render_template("General_Form.html", backgroud_Filename=background_pic('upload_Swaps_csv'),
                           form=form, Table_name="Swaps", header=header, description=description, title=title, no_backgroud_Cover=True)


@swaps.route('/Swaps/Other_Brokers', methods=['GET', 'POST'])
@roles_required(["Risk", "Risk_TW","Risk_UK", "Admin"])
def Other_Brokers():

    title = "Swap Compare"
    header = "Swap Compare"

    description = Markup("Swaps from other brokers.<br>" +
                         "fxdd: fxdd (https://www.fxdd.com/mt/en/trading/offering)<br>" +
                         "fdc: Forex.com (https://www.forex.com/en/trading/pricing-fees/rollover-rates/)<br>" +
                         "tv: Trade-View(https://www.tradeviewforex.com/room/forex-resources/rollover-rates)<br>" +
                         "gp: Global Prime (https://www.globalprime.com/trading-conditions/swaps-financing/)<br>" +
                         "saxo: saxo (https://www.home.saxo/en-sg/rates-and-conditions/forex/trading-conditions#historic-swap-points)<br>" +
                         "ebh: European Brokerage House (https://ebhforex.com/faq/rollover-policy/)<br>" +
                         "fpm: fpmarkets (https://www.fpmarkets.com/swap-point)<br>"+
                         "cfh: FRom CFH back office. Will divide by the number of days of swaps (ie: not showing 3 days worth..)<br>" +
                        "For CFH Swaps, CFH dosn't use digits for CFD calculations. So we need to multiply it by the digits for CFD. <br>" +
                        "CFD also has dividend included.<br>" +
                        "CFH - For FX, their value are in Pips, but since we upload in points, we need to X10.<br>" +
                        "<br> CFH Swaps has been inverted to reflect the same as MT4. Below are from the original instruction, if taken from their Back office.<br>" +
                        "CFH -  if the swap rate is positive for the short side, BGI will receive swap. <br>" +
                        "CFH - if positive for the long side, you will pay swap and vice versa.")


        # TODO: Add Form to add login/Live/limit into the exclude table.
    return render_template("Webworker_Single_Table_No_Border.html", backgroud_Filename='css/Person_Mac.jpg', icon="",
                           Table_name="Swap Compare ", title=title,
                            ajax_url=url_for('swaps.Other_Brokers_Ajax', _external=True), header=header,
                           description=description, replace_words=Markup(["Today"]))


@swaps.route('/Swaps/Other_Brokers_Ajax', methods=['GET', 'POST'])
@roles_required(["Risk", "Risk_TW","Risk_UK", "Admin", "Dealing"])
def Other_Brokers_Ajax():

    start = datetime.datetime.now()
    # Get swaps from other brokers
    df_other_broker_swaps = get_broker_swaps(db)

    # Want to get BGI Swaps to do a left join with.
    sql_query_line = """select Core_Symbol as Symbol
    FROM aaron.swap_bgicoresymbol
    ORDER BY Symbol"""

    df_bgi_core_symbol = get_from_sql_or_file(sql_query_line, current_app.config["VANTAGE_UPLOAD_FOLDER"] + "BGI_Core_Symbol_Only.xls", db)


    df_other_broker_swaps = df_bgi_core_symbol.merge(df_other_broker_swaps, on="Symbol", how="left")

    df_other_broker_swaps.fillna("-", inplace=True)

    #return json.dumps(pd_dataframe_to_dict(df_fdc))
    print("Swap Compare took: {}s".format((datetime.datetime.now() - start).total_seconds()))
    return json.dumps(pd_dataframe_to_dict(df_other_broker_swaps))



# Want to get the digits changes and get the upload file to OZ
@swaps.route('/Swaps/CFH_OZ_Upload')
@roles_required()
def cfh_oz_upload():

    description = Markup("Taking Values off CFH Backoffice Via SOAP.<br>Digits compensation for OZ-CFH has been done.")
    # Template will take in an json dict and display a few tables, depending on how many is needed.
    return render_template("Webworker_1_table_Boderless_excel.html",
                           backgroud_Filename='css/Color-Pencil.jpg', \
                           title="CFH OZ Swaps",
                           header="CFH OZ Swaps",
                           excel_file_name="OZ_Swap_Upload.xlsx",
                           ajax_url=url_for('swaps.cfh_oz_upload_Ajax', _external=True),
                           description=description, replace_words=Markup(["Today"]))


# The Ajax to get the OZ file.
@swaps.route('/Swaps/CFH_OZ_Upload_Ajax', methods=['GET', 'POST'])
def cfh_oz_upload_Ajax():

    df_cfd_conversion = get_OZ_CFH_cfd_Digits()

    # This is using unsync. Need to query .result() to get the return data.
    # Does the wait when calling .result()
    cfh_oz_swaps = CFH_Soap_Swaps(backtrace_days_max=1, divide_by_days=False, cfd_conversion=False, df_cfd_conversion=df_cfd_conversion)

    return json.dumps({"Swaps": cfh_oz_swaps.result()})
    #return json.dumps([{"Testing": "12345"}])


def markup_swaps(Val, positive_markup, negative_markup ):
    val = Val
    if val >= 0:  # Positive!
        markup_percentage = float(positive_markup)
        val = val * (100 - markup_percentage) / 100
    else:
        markup_percentage = float(negative_markup)
        val = val * (100 + markup_percentage) / 100
    return val


# To view Client's trades as well as some simple details.
@swaps.route('/Swap_upload_form', methods=['GET', 'POST'])
@roles_required()
def Swap_upload_form():

    test = False

    title = "Swap Upload"
    header = "Swap Upload"
    description = Markup("Swap Upload.")
    # file_form = File_Form()
    # and form.validate_on_submit()




    #print(request.method)
    today_swap_SQL_Query = """Select `core_symbol` as `Core Symbol`, `vantage_long` as `Long Points`, `vantage_short` as `Short Points`
                          FROM aaron.swaps_vantage_raw
                           WHERE date='{}'""".format(datetime.datetime.now().strftime("%Y-%m-%d"))


    # today_swap_SQL_Query = """Select `core_symbol` as `Core Symbol`, `vantage_long` as `Long Points`, `vantage_short` as `Short Points`
    #                       FROM aaron.swaps_vantage_raw
    #                        WHERE date='2020-01-01' """

    file_data = query_SQL_return_record(today_swap_SQL_Query)   # Get the uploaded swaps from SQL.

    # Check that we got data from SQL.
    if len(file_data) == 0:
        if test:
        # Artificially create something.
            file_data = [{'Core Symbol': 'MXN/JPY', 'Long Points': -0.29, 'Short Points': -1.46, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'HKD/JPY', 'Long Points': 0.15, 'Short Points': -0.25, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'NOK/JPY', 'Long Points': 0.39, 'Short Points': -0.54, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'GBP/SEK', 'Long Points': 1.82, 'Short Points': -3.44, 'Symbol Group Path': '*', 'digits': 4}, {'Core Symbol': 'SEK/JPY', 'Long Points': 0.2, 'Short Points': -0.14, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'NOK/SEK', 'Long Points': 1.15, 'Short Points': -1.96, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'EUR/DKK', 'Long Points': 12.13, 'Short Points': -9.03, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'EUR/SEK', 'Long Points': 27.31, 'Short Points': -1.62, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'CAD/SGD', 'Long Points': -0.03, 'Short Points': -1.77, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'GBP/NOK', 'Long Points': 2.2, 'Short Points': -2.69, 'Symbol Group Path': '*', 'digits': 4}, {'Core Symbol': 'EUR/PLN', 'Long Points': 26.07, 'Short Points': -9.57, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'EUR/NOK', 'Long Points': 31.52, 'Short Points': 1.63, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'GBP/HKD', 'Long Points': 21.47, 'Short Points': -24.12, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'EUR/HKD', 'Long Points': 29.07, 'Short Points': 2.91, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'EUR/TRY', 'Long Points': 640.42, 'Short Points': 0.0, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'USD/MXN', 'Long Points': 342.23, 'Short Points': 0.0, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'USD/TRY', 'Long Points': 515.32, 'Short Points': 0.0, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'EUR/MXN', 'Long Points': 453.16, 'Short Points': 0.0, 'Symbol Group Path': '*', 'digits': 5}, \
                         {'Core Symbol': 'USD/RUB', 'Long Points': 13.11, 'Short Points': 4.63, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'USD/CZK', 'Long Points': 0.22, 'Short Points': 0.02, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'USD/DKK', 'Long Points': -3.18, 'Short Points': -15.72, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'USD/HKD', 'Long Points': -0.3, 'Short Points': -5.79, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'USD/HUF', 'Long Points': 11.73, 'Short Points': -5.37, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'USD/PLN', 'Long Points': 12.24, 'Short Points': -16.07, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'USD/ZAR', 'Long Points': 219.54, 'Short Points': 0.0, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'USD/SEK', 'Long Points': 4.58, 'Short Points': -20.91, 'Symbol Group Path': '*', 'digits': 5}, \
                         {'Core Symbol': 'EUR/GBP', 'Long Points': 3.66, 'Short Points': -0.36, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'USD/CNH', 'Long Points': 70.44, 'Short Points': 18.59, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'USD/CHF', 'Long Points': -0.94, 'Short Points': -4.05, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'USD/SGD', 'Long Points': 2.75, 'Short Points': -2.75, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'EUR/JPY', 'Long Points': 3.65, 'Short Points': 0.1, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'EUR/USD', 'Long Points': 4.57, 'Short Points': 0.68, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'EUR/NZD', 'Long Points': 6.52, 'Short Points': 1.27, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'USD/NOK', 'Long Points': 7.72, 'Short Points': -16.1, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'EUR/SGD', 'Long Points': 6.7, 'Short Points': 0.12, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'USD/CAD', 'Long Points': 2.21, 'Short Points': -1.88, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'USD/JPY', 'Long Points': 1.3, 'Short Points': -2.91, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'GBP/USD', 'Long Points': 3.53, 'Short Points': -2.75, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'GBP/AUD', 'Long Points': 3.32, 'Short Points': -4.63, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'GBP/CAD', 'Long Points': 4.42, 'Short Points': -3.04, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'GBP/CHF', 'Long Points': -0.37, 'Short Points': -6.24, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'GBP/SGD', 'Long Points': 4.14, 'Short Points': -3.2, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'GBP/NZD', 'Long Points': 5.28, 'Short Points': -3.53, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'GBP/JPY', 'Long Points': 2.45, 'Short Points': -3.92, 'Symbol Group Path': '*', 'digits': 3}, \
                         {'Core Symbol': 'CHF/JPY', 'Long Points': 4.17, 'Short Points': 0.77, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'AUD/NZD', 'Long Points': 2.44, 'Short Points': -0.64, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'AUD/SGD', 'Long Points': 2.63, 'Short Points': -1.5, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'AUD/CAD', 'Long Points': 1.99, 'Short Points': -0.61, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'CAD/JPY', 'Long Points': 0.43, 'Short Points': -1.96, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'EUR/AUD', 'Long Points': 4.69, 'Short Points': 0.27, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'EUR/CHF', 'Long Points': 0.78, 'Short Points': -2.36, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'AUD/JPY', 'Long Points': 1.0, 'Short Points': -1.23, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'CAD/CHF', 'Long Points': -0.86, 'Short Points': -3.3, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'AUD/CHF', 'Long Points': -0.42, 'Short Points': -2.59, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'EUR/CAD', 'Long Points': 5.5, 'Short Points': 1.04, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'NZD/SGD', 'Long Points': 1.72, 'Short Points': -2.17, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'NZD/USD', 'Long Points': 1.18, 'Short Points': -1.55, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'AUD/USD', 'Long Points': 2.46, 'Short Points': -1.54, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'NZD/CAD', 'Long Points': 1.3, 'Short Points': -1.3, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'NZD/CHF', 'Long Points': -0.78, 'Short Points': -2.97, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'NZD/JPY', 'Long Points': 0.32, 'Short Points': -1.74, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'SGD/JPY', 'Long Points': 1.04, 'Short Points': -2.23, 'Symbol Group Path': '*', 'digits': 3}, \
                         {'Core Symbol': 'XAU/USD', 'Long Points': 3.69, 'Short Points': -1.18, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'XAG/USD', 'Long Points': 0.6, 'Short Points': -0.2, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'XPT/USD', 'Long Points': 13.69, 'Short Points': -3.5, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'XPD/USD', 'Long Points': 32.56, 'Short Points': -8.99, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'XAU/AUD', 'Long Points': 0.33, 'Short Points': -0.32, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'XAG/AUD', 'Long Points': 0.6, 'Short Points': -0.45, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'CHF/SGD', 'Long Points': 9.14, 'Short Points': -2.54, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'AUD/CNH', 'Long Points': 84.85, 'Short Points': 10.56, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'BCHUSD', 'Long Points': 28.29, 'Short Points': -28.29, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'LTCUSD', 'Long Points': 7.89, 'Short Points': -7.89, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'ETHUSD', 'Long Points': 121.5, 'Short Points': -121.5, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'XRPUSD', 'Long Points': 3.81, 'Short Points': -3.81, 'Symbol Group Path': '*', 'digits': 4}, {'Core Symbol': 'BTCUSD', 'Long Points': 2033.02, 'Short Points': -2033.02, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'COPPER-C', 'Long Points': 4.63, 'Short Points': -2.1, 'Symbol Group Path': '*', 'digits': 4}, {'Core Symbol': 'NG-C', 'Long Points': 1.84, 'Short Points': 0.44, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'GAS-C', 'Long Points': 3.22, 'Short Points': -0.17, 'Symbol Group Path': '*', 'digits': 4}, {'Core Symbol': 'Cocoa-C', 'Long Points': 6.06, 'Short Points': 1.77, 'Symbol Group Path': '*', 'digits': 1}, {'Core Symbol': 'Coffee-C', 'Long Points': 5.28, 'Short Points': 1.99, 'Symbol Group Path': '*', 'digits': 4}, \
                         {'Core Symbol': 'Cotton-C', 'Long Points': 16.73, 'Short Points': 2.11, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'OJ-C', 'Long Points': 7.28, 'Short Points': 3.74, 'Symbol Group Path': '*', 'digits': 4}, {'Core Symbol': 'Sugar-C', 'Long Points': 1.46, 'Short Points': -1.3, 'Symbol Group Path': '*', 'digits': 5}, {'Core Symbol': 'GASOIL-C', 'Long Points': 9.86, 'Short Points': 0.32, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'Soybean-C', 'Long Points': -8.72, 'Short Points': -14.62, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'Wheat-C', 'Long Points': 1.46, 'Short Points': 0.25, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'CHINA50', 'Long Points': 0.85, 'Short Points': -0.77, 'Symbol Group Path': '*', 'digits': 0}, {'Core Symbol': 'DAX30', 'Long Points': 67.24, 'Short Points': -133.11, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'NAS100', 'Long Points': 101.64, 'Short Points': -91.8, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'Nikkei225', 'Long Points': 173.96, 'Short Points': -189.64, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'HSI', 'Long Points': 1.84, 'Short Points': -1.69, 'Symbol Group Path': '*', 'digits': 0}, {'Core Symbol': 'SPI200', 'Long Points': 48.4, 'Short Points': -47.31, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'FTSE100', 'Long Points': -83.61, 'Short Points': -172.91, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'EU50', 'Long Points': 17.48, 'Short Points': -34.6, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'SP500', 'Long Points': -3.45, 'Short Points': -60.26, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'DJ30', 'Long Points': -316.75, 'Short Points': -803.88, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'FRA40', 'Long Points': 28.08, 'Short Points': -55.58, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'ES35', 'Long Points': 52.75, 'Short Points': -83.06, 'Symbol Group Path': '*', 'digits': 2}, \
                         {'Core Symbol': 'ITA40', 'Long Points': 152.91, 'Short Points': -240.78, 'Symbol Group Path': '*', 'digits': 2}, {'Core Symbol': 'UKOUSD', 'Long Points': -14.05, 'Short Points': -32.41, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': 'USOUSD', 'Long Points': -3.55, 'Short Points': -16.32, 'Symbol Group Path': '*', 'digits': 3}, {'Core Symbol': '*', 'Long Points': 0, 'Short Points': 0, 'Symbol Group Path': '*', 'digits': 0}]

        else:
            return redirect(url_for('swaps.upload_Swaps_csv'))  # Go to the swap upload page
        return redirect(url_for('swaps.upload_Swaps_csv'))   # Go to the Swap Upload Page

    if request.method != 'POST':

        df_swap_data = calculate_swaps_bgi(file_data, db) # Get the data processed by the helper function
        #print(df)
        df_swap_data.fillna("-", inplace=True) # Fill the NAs so that it will not appear weird.
        data = df_swap_data.to_dict("records")

        form = All_Swap_Form()
        # Live = form.Live.data  # Get the Data.
        # Login = form.Login.data
        #form.title.data = "All Swaps"  # change the field's data

        for f in data:  # Loop thru all swaps uploaded.
            if all([u in f for u in ['bgi_coresymbol', 'long_markup_value_PlusFixed', 'short_markup_value_PlusFixed',
                        'swap_markup_profile', "long_markup_value_Plus_Insti_Fixed",
                        "short_markup_value_Plus_Insti_Fixed",
                        'avg_long',  'avg_short', 'tv Long',
                        "dividend", "Markup_Style",
                        'tv Short',  'gp Long', 'gp Short']]):

                symbol_form = Individual_symbol_Form()
                symbol_form.symbol = f["bgi_coresymbol"]
                symbol_form.long = f["long_markup_value_PlusFixed"]
                symbol_form.short = f["short_markup_value_PlusFixed"]

                # To keep track if the Long and Short data has been changed.
                symbol_form.Long_Hidden = f["long_markup_value_PlusFixed"]
                symbol_form.Short_Hidden = f["short_markup_value_PlusFixed"]


                symbol_form.avg_short = f["avg_short"]
                symbol_form.avg_long = f["avg_long"]

                symbol_form.broker_1_long = f["gp Long"]
                symbol_form.broker_1_short = f["gp Short"]

                symbol_form.broker_2_long = f["tv Long"]
                symbol_form.broker_2_short = f["tv Short"]


                symbol_form.bloomberg_dividend = f["dividend"]

                symbol_form.symbol_markup_type = f["swap_markup_profile"]
                symbol_form.symbol_markup_style = f["Markup_Style"]

                symbol_form.insti_long = f["long_markup_value_Plus_Insti_Fixed"]
                symbol_form.insti_short  = f["short_markup_value_Plus_Insti_Fixed"]

                # The cell color
                symbol_form.long_style = compare_swap_values(f["long_markup_value_PlusFixed"], f["avg_long"])
                symbol_form.short_style = compare_swap_values(f["short_markup_value_PlusFixed"], f["avg_short"])

                symbol_form.broker_1_long_style = compare_swap_values(f["long_markup_value_PlusFixed"], f["gp Long"])
                symbol_form.broker_1_short_style = compare_swap_values(f["short_markup_value_PlusFixed"], f["gp Short"])

                symbol_form.broker_2_long_style = compare_swap_values(f["long_markup_value_PlusFixed"], f["tv Long"])
                symbol_form.broker_2_short_style = compare_swap_values(f["short_markup_value_PlusFixed"], f["tv Short"])

                # Append to the Swaps for all.
                form.core_symbols.append_entry(symbol_form)
            else:
                pass


    if request.method == 'POST':
        form = All_Swap_Form()


        if form.validate_on_submit():

            all_data = [[s.symbol.data, s.long.data, s.short.data, s.insti_long.data, s.insti_short.data,
                                        s.Long_Hidden.data, s.Short_Hidden.data ] for s in form.core_symbols ]
            # for s in form.core_symbols:
            #     # Append to the list.
            #     all_data.append([s.symbol.data, s.long.data, s.short.data])
            # for s in all_data:
            #     print(s)
            #print("{} | {} | {} | {} | {}".format(s.symbol.data, s.long.data, s.short.data, s.insti_long.data, s.insti_short.data))

            process_validated_swaps(all_data) # Get it inserted into SQL and run upload to MT4/5
            return redirect(url_for('swaps.Swap_download_page'))

        else:
            flash("Kindly Check values. Can't submit due to error in the swap numbers.", 'error')
            print("Can't validate Swap results.")

    return render_template("Swap_Calculate_results.html",
                           title=title, backgroud_Filename = background_pic("Swap_upload_form"),
                           header=header,
                           form=form, no_backgroud_Cover=True,
                           description=description)


@swaps.route('/Swap_download', methods=['GET', 'POST'])
@roles_required()
def Swap_download_page():


    title = "Swap Upload File"
    header = "Swap Upload File"
    description = Markup("Swap Upload File")


    return render_template("Swap_Calculate_results.html",
                           title=title, show_table=True,
                           header=header, backgroud_Filename = background_pic("Swap_download_page"),
                           no_backgroud_Cover=True,
                           description=description)


# #To Let the user download the Swap file, IF it's in the session's cookie.
# # To view Client's trades as well as some simple details.
@swaps.route('/Swap_download_excel', methods=['GET', 'POST'])
@roles_required()
def Swap_download_excel():


    # Get it from SQL.
    today_swap_SQL_Query = """SELECT S.Core_Symbol AS `Core Symbol (BGI)`, `Long Points (BGI)`, `Short Points (BGI)`,
        COALESCE(swap_bgifixedswap_insti.`long`,  `Long Points (BGI)`) AS `Insti Long Points (BGI)`, 
        COALESCE(swap_bgifixedswap_insti.`short`,  `Short Points (BGI)`) AS `Insti Short Points (BGI)`
        FROM	
        (SELECT Core_Symbol, 
        BGI_LONG AS `Long Points (BGI)`, 
        BGI_SHORT AS `Short Points (BGI)`         
        FROM aaron.`bgi_swaps` AS s
        WHERE  s.DATE = '{}'
        AND s.Core_Symbol not like "*") AS S
        LEFT JOIN
        aaron.swap_bgifixedswap_insti ON swap_bgifixedswap_insti.core_symbol = S.Core_Symbol
        """.format(datetime.datetime.now().strftime("%Y-%m-%d"))

    swap_data = query_SQL_return_record(today_swap_SQL_Query)   # Get the uploaded swaps from SQL.

    # Check that there are returns. Else, remove it all.
    if len(swap_data) == 0:
        print("swap_validated_data NOT in SQL. ")
        return redirect(url_for('swaps.upload_Swaps_csv'))

    #Make the pandas dataframe.
    df = pd.DataFrame(swap_data, columns=["Core Symbol (BGI)", "Long Points (BGI)", "Short Points (BGI)", "Insti Long Points (BGI)", "Insti Short Points (BGI)"])

    #print(df)

    # Need to cast to float so that it would be saved as "number" in excel.
    for c in ["Long Points (BGI)", "Short Points (BGI)", "Insti Long Points (BGI)", "Insti Short Points (BGI)"]:
        df[c] = df[c].astype(float)


    df.sort_values("Core Symbol (BGI)", inplace=True)


    retail_sheet_col = ["Core Symbol (BGI)", "Long Points (BGI)", "Short Points (BGI)"]
    insti_sheet_col = ["Core Symbol (BGI)", "Insti Long Points (BGI)", "Insti Short Points (BGI)"]

    # ---- Start Drawing the Excel Style.
    # Let's create a style template for the header row

    #print(df[retail_sheet_col])

    workbook = Workbook()
    sheet = workbook.active

    # ---- Start Drawing the Excel Style.
    # Let's create a style template for the header row
    header = NamedStyle("Header")
    header.font = Font(bold=True)
    header.fill = PatternFill("solid", fgColor="00C7E2FF")
    thin = Side(border_style="thin", color="000000")
    header.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header.alignment = Alignment(horizontal="center", vertical="center")

    insti_sheet = workbook.create_sheet("new")    # Insti Sheet

    # Two Tabs. The functions will overwrite the sheets they are given
    # Need to rename the columns
    df_insti = df[insti_sheet_col].rename(columns={"Insti Long Points (BGI)" : "Long Points (BGI)",
                                                   "Insti Short Points (BGI)": "Short Points (BGI)"})

    generate_pretty_Swap_file(df[retail_sheet_col], sheet, "retail", header_style=header)
    generate_pretty_Swap_file(df_insti, insti_sheet, "insti", header_style=header)


    content = save_virtual_workbook(workbook)
    resp = make_response(content)
    resp.headers["Content-Disposition"] = 'attachment; filename=MT4Swaps {dt.day} {dt:%b} {dt.year}.xls'.format(dt=datetime.datetime.now())
    #resp.headers['Content-Type'] = 'application/x-xls'
    resp.headers["Content-type"] = "application/vnd.ms-excel"


    return resp




@swaps.route('/add_swap_markup_profile', methods=['GET', 'POST'])      # Want to add an offset to the ABook page.
@roles_required()
def add_swap_markup_profile():

    title = "Add Swap Markup Profile"
    header = "Add Swap Markup Profile"
    description = Markup("Adding Swap markup profile. <br>SQL Table: <b>aaron.swap_markup_profile</b><br>" +\
                         "Updates will <b>cover/overwrite</b> the current settings, if markup names are duplicated.")

    form = AddMarkupProfile()
    if request.method == 'POST' and form.validate_on_submit():
        #pass
        Markup_Profile = form.Markup_Profile.data       # Get the Data.
        Long_Markup = form.Long_Markup.data
        Short_Markup = form.Short_Markup.data

        print(Markup_Profile)
        sql_insert = "INSERT INTO  aaron.`Swap_markup_Profile` (`Swap_markup_profile`, `Long_Markup`, `Short_Markup`) VALUES" \
            " ('{}','{}','{}' ) ON DUPLICATE KEY UPDATE `Long_Markup` = VALUES(`Long_Markup`), `Short_Markup` = VALUES(`Short_Markup`)".format(Markup_Profile, Long_Markup, Short_Markup)
        print(sql_insert)
        db.engine.execute(text(sql_insert))   # Insert into DB
        flash(Markup("<b>{Markup_Profile}</b> has been updated the database.".format(Markup_Profile=Markup_Profile)))


    # For FLASK-TABLE to work. We need to get the names from SQL right.
    # Also, we want to get the Symbols to upper for better display. That's all. ha ha


    sql_query = "SELECT * FROM swap_markup_profile ORDER BY Swap_markup_profile DESC"
    collate = query_SQL_return_record(text(sql_query))

    if len(collate) == 0:   # There is no data.
        empty_table = [{"Result": "There are currently no profile"}]
        table = create_table_fun(empty_table, additional_class=["basic_table", "table", "table-striped",
                                                                "table-bordered", "table-hover", "table-sm"])
    else:
        # Live Account and other information that would be going into the table.
        df_data = pd.DataFrame(collate)
        table = Delete_Swap_Profile_Table(df_data.to_dict("record"))
        table.html_attrs = {"class": "basic_table table table-striped table-bordered table-hover table-sm"}


    return render_template("General_Form.html", form=form, table=table, title=title,
                           header=header,description=description,
                           backgroud_Filename=background_pic( "add_swap_markup_profile" ))


# To remove the swap profile from our server.
@swaps.route('/Remove_Swap_Markup_Profile/<Swap_Profile>', methods=['GET', 'POST'])
@roles_required()
def Remove_Swap_Markup_Profile_Endpoint(Swap_Profile=""):


    # First we want to check if there are any symbols using the current swap profile that we want to delete.
    # If there is, we will not allow it to be deleted.
    sql_query = """select Core_Symbol 
                From aaron.swap_bgicoresymbol 
                    WHERE Swap_Markup_Profile = '{}'""".format(Swap_Profile)
    raw_result = db.engine.execute(text(sql_query))
    result_data = raw_result.fetchall()
    # result_col = raw_result.keys()
    # print(len(result_data))
    # print(result_data)

    if len(result_data) > 0:    # There are Symbols using this swap markup profile. Do not allow deletion.
        symbols_using_profile = [r[0] for r in result_data]
        symbol_string = ",".join(symbols_using_profile[:5]) # Take at most 5..
        flash(Markup("<b>{Swap_Profile}</b> Markup Profile <b>Cannot be deleted.</b><br>There are symbols using that markup profile: <b>{symbol_string}</b>".format(Swap_Profile=Swap_Profile,
                                                                                                                                                   symbol_string=symbol_string)), "error")
        return redirect(url_for('swaps.add_swap_markup_profile'))


    # # # Write the SQL Statement to write the values into SQL to clear the offset, By Symbols.
    # # # It will write -1 * consolidated value into SQL.
    sql_insert_statement = """DELETE FROM aaron.swap_markup_profile WHERE swap_markup_profile = '{Swap_Profile}'""".format(Swap_Profile=Swap_Profile)

    sql_update_statement = sql_insert_statement.replace("\n", "").replace("\t", "")
    # # print(sql_update_statement)
    sql_update_statement = text(sql_update_statement)
    print(sql_insert_statement)
    result = db.engine.execute(sql_update_statement)

    flash(Markup("<b>{Swap_Profile}</b> Markup Profile has been deleted".format(Swap_Profile=Swap_Profile)))
    return redirect(url_for('swaps.add_swap_markup_profile'))





# Want to check and close off account/trades.
@swaps.route('/Symbol_Markup_Profile/Settings', methods=['GET', 'POST'])
@roles_required()
def Symbol_Markup_Profile_Settings():



    title = Markup("Symbol Markup Profile Settings")
    header = "Symbol Markup Profile"
    description = Markup("Symbol Markup Profile Settings.<br>The updated swap markup profile will overwrite the current markup profile.<br>" + \
                         "BGI Core symbols SQL Table: aaron.swap_bgicoresymbol<br>" +\
                         "Markup Profile SQL Table: aaron.swap_markup_profile<br>")

    form = Symbol_Swap_Profile_Form()

    if request.method == 'POST' and form.validate_on_submit():

        Symbol = form.Symbol.data
        Markup_Profile = form.Markup_Profile.data

        sql_insert = """UPDATE aaron.swap_bgicoresymbol SET `swap_markup_profile` = '{Markup_Profile}' WHERE `Core_Symbol` = '{Symbol}' """.format(
                Symbol=Symbol, Markup_Profile=Markup_Profile)

        sql_insert = sql_insert.replace("\t", "").replace("\n", "")

        print(sql_insert)
        db.engine.execute(text(sql_insert))  # Insert into DB
        flash("Symbol: {} | Markup Profile: {}".format(Symbol, Markup_Profile))


    # Want to select all the Core Symbol
    sql_query = """select Core_Symbol, Contract_Size, Digits, Currency, Swap_Markup_Profile From aaron.swap_bgicoresymbol"""
    raw_result = db.engine.execute(sql_query)
    result_data = raw_result.fetchall()
    result_col = raw_result.keys()
    #print(result_data)
    Symbol_List = [(r[0], "{} | [{}]".format(r[0], r[4])) for r in result_data if len(r)>=5]

    # Want to select all the markup profile
    sql_query_mup = """select swap_markup_profile from aaron.swap_markup_profile"""
    raw_result_mup = db.engine.execute(sql_query_mup)
    result_data_mup = raw_result_mup.fetchall()
    #print(result_data_mup)
    markup_profile_list = [(r[0], r[0]) for r in result_data_mup if len(r)>0]


    # passing Symbol_List and markup_profile_list to the form
    form.Symbol.choices = Symbol_List
    form.Markup_Profile.choices = markup_profile_list



    # Want to create the table to output data with the data from the SQL query above.
    if len(result_data) == 0:   # There is no data.
        collate = [{"Result": "There are currently no Accounts being monitored"}]
        table = create_table_fun(collate, additional_class=["basic_table", "table", "table-striped", "table-bordered", "table-hover", "table-sm"])
    else:
        # Output all of BGI Core symbol with their respective data.
        bgi_core_symbol = [dict(zip(result_col, r)) for r in result_data]
        #print(bgi_core_symbol)
        table = Symbol_Swap_Profile_Table(bgi_core_symbol)
        table.html_attrs = {"class": "basic_table table table-striped table-bordered table-hover table-sm"}

    #return table


    # flash("{symbol} {offset} updated in A Book offset.".format(symbol=symbol, offset=offset))
    # backgroud_Filename='css/Equity_cut.jpg', Table_name="Equity Protect Cut",  replace_words=Markup(["Today"])
    # TODO: Add Form to add login/Live/limit into the exclude table.
    return render_template("General_Form.html",
                           title=title, header=header,
                           form=form, description=description, table=table,
                           backgroud_Filename=background_pic("Symbol_Markup_Profile_Settings"))


# # # NOT USING as this was the excel with no formatting.
# #To Let the user download the Swap file, IF it's in the session's cookie.
# # To view Client's trades as well as some simple details.
# @swaps.route('/Swap_download_excel_2', methods=['GET', 'POST'])
# @roles_required()
# def Swap_download_excel_2():
#
#     # Get it from SQL.
#     today_swap_SQL_Query = """SELECT S.Core_Symbol AS `Core Symbol (BGI)`, `Long Points (BGI)`, `Short Points (BGI)`,
#         COALESCE(swap_bgifixedswap_insti.`long`,  `Long Points (BGI)`) AS `Insti Long Points (BGI)`,
#         COALESCE(swap_bgifixedswap_insti.`short`,  `Short Points (BGI)`) AS `Insti Short Points (BGI)`
#         FROM
#         (SELECT Core_Symbol,
#         BGI_LONG AS `Long Points (BGI)`,
#         BGI_SHORT AS `Short Points (BGI)`
#         FROM aaron.`bgi_swaps` AS s
#         WHERE  s.DATE = '{}'
#         AND s.Core_Symbol not like "*") AS S
#         LEFT JOIN
#         aaron.swap_bgifixedswap_insti ON swap_bgifixedswap_insti.core_symbol = S.Core_Symbol
#         """.format(datetime.datetime.now().strftime("%Y-%m-%d"))
#
#     swap_data = query_SQL_return_record(today_swap_SQL_Query)   # Get the uploaded swaps from SQL.
#
#     # Check that there are returns. Else, remove it all.
#     if len(swap_data) == 0:
#         print("swap_validated_data NOT in SQL. ")
#         return redirect(url_for('swaps.upload_Swaps_csv'))
#
#     #Make the pandas dataframe.
#     df = pd.DataFrame(swap_data, columns=["Core Symbol (BGI)", "Long Points (BGI)", "Short Points (BGI)", "Insti Long Points (BGI)", "Insti Short Points (BGI)"])
#
#     # Need to cast to float so that it would be saved as "number" in excel.
#     for c in ["Long Points (BGI)", "Short Points (BGI)", "Insti Long Points (BGI)", "Insti Short Points (BGI)"]:
#         df[c] = df[c].astype(float)
#
#
#     df.sort_values("Core Symbol (BGI)", inplace=True)
#
#
#     retail_sheet = [["Core Symbol (BGI)",	"Long Points (BGI)", "Short Points (BGI)"]] + df[["Core Symbol (BGI)", "Long Points (BGI)", "Short Points (BGI)"]].values.tolist()
#     insti_sheet = [["Core Symbol (BGI)",	"Long Points (BGI)", "Short Points (BGI)"]] + df[["Core Symbol (BGI)", "Insti Long Points (BGI)", "Insti Short Points (BGI)"]].values.tolist()
#
#     content = {'retail': retail_sheet,
#                'insti': insti_sheet, }
#
#     book = pyexcel.Book(content)
#
#     # pip install pyexcel-xls
#     return excel.make_response(book, file_type="xls", file_name='MT4Swaps {dt.day} {dt:%b} {dt.year}'.format(dt=datetime.datetime.now()))



#
#
# @swaps.route("/upload", methods=['GET', 'POST'])
# def upload_file():
#     if request.method == 'POST':
#         return jsonify({"result": request.get_array(field_name='file')})
#     return '''
#     <!doctype html>
#     <title>Upload an excel file</title>
#     <h1>Excel file upload (csv, tsv, csvz, tsvz only)</h1>
#     <form action="" method=post enctype=multipart/form-data><p>
#     <input type=file name=file><input type=submit value=Upload>
#     </form>
#     '''




# @swaps.route("/custom_export", methods=['GET'])
# def docustomexport():
#
#     # wb = Workbook()
#     # ws1 = wb.create_sheet("Sheet_A")
#     # ws1.title = "Title_A"
#     #
#     # ws2 = wb.create_sheet("Sheet_B", 0)
#     # ws2.title = "Title_B"
#
#     # Need to do this pip install
#     # pip install pyexcel-xlsx
#
#     content = { 'Sheet 1': [[1.0, 2.0, 3.0]],
#                 'Sheet 2': [ ['X', 'Y', 'Z'], [1.0, 2.0, 3.0], [4.0, 5.0, 6.0] ],
#                 'Sheet 3': [  ['O', 'P', 'Q'],  [3.0, 2.0, 1.0], [4.0, 3.0, 2.0] ],
#                 }
#
#     book = pyexcel.Book(content)
#
#     return excel.make_response(book, file_type ="xls", file_name="export_data")
#     #return excel.make_response_from_array([[1, 2], [3, 4]], "xlsx", file_name="export_data")



#
# # Want to insert into table.
# # From Flask.
# @swaps.route('/Upload_Swap_File', methods=['GET', 'POST'])
# @roles_required()
# def Upload_Swap_File():
#     title = Markup("Upload Swap File")
#     header = title
#     description = Markup("Upload Swap file for calculations")
#
#     form = UploadForm()
#     print("Method: {}".format(request.method))
#     print("validate_on_submit: {}".format(form.validate_on_submit()))
#     form.validate_on_submit()
#
#     if request.method == 'POST' and form.validate_on_submit():
#         filename = secure_filename(form.file.data.filename)
#         print(form.file.data)
#
#
#         #flash("Live: {live}, Login: {login} Equity limit: {equity_limit} has been added to live1.`balance_equity_exclude`.".format(live=Live, login=Login, equity_limit=Equity_Limit))
#
#     # flash("{symbol} {offset} updated in A Book offset.".format(symbol=symbol, offset=offset))
#     # backgroud_Filename='css/Equity_cut.jpg', Table_name="Equity Protect Cut",  replace_words=Markup(["Today"])
#     # TODO: Add Form to add login/Live/limit into the exclude table.
#     return render_template("General_Form.html", backgroud_Filename=background_pic("Upload_Swap_File"),
#                            title=title, header=header,
#                            form=form, description=description)





# @app.route('/upload')
# def upload_file2():
#  return '''
# <html>
# <body>
#    <form action = "http://localhost:5000/uploader" method = "POST"
#       enctype = "multipart/form-data">
#       <input type = "file" name = "file" />
#       <input type = "submit"/>
#    </form>
# </body>
# </html>
#  '''

#
# @app.route('/uploader', methods=['GET', 'POST'])
# def upload_file():
#  if request.method == 'POST':
#      f = request.files['file']
#      f.save(secure_filename(f.filename))
#      return 'file uploaded successfully'
#
#
# @app.route('/uploads/<filename>')
# def uploaded_file(filename):
#  return send_from_directory(app.config['UPLOAD_FOLDER'],
#                             filename)



#
# @app.route('/test1', methods=['GET', 'POST'])
# def retrieve_db_swaps():
#  raw_result = db.engine.execute("select * from aaron.swap_bgi_vantage_coresymbol")
#  result_data = raw_result.fetchall()
#  result_col = raw_result.keys()
#  result_colate = [dict(zip(result_col, a)) for a in result_data]
#
#  T = create_table()
#  table = T(result_colate, classes=["table", "table-striped", "table-bordered", "table-hover"])
#  if (len(result_colate) > 0) and isinstance(result_colate[0], dict):
#      for c in result_colate[0]:
#          if c != "\n":
#              table.add_column(c, Col(c, th_html_attrs={"style": "background-color:# afcdff"}))
#  return render_template("Swap_Sql.html", table=table)

#
# @app.route('/upload_swap', methods=['GET', 'POST'])
# def retrieve_db_swaps2():
#  form = UploadForm()
#
#  if request.method == 'POST' and form.validate_on_submit():
#
#      record_dict = request.get_records(field_name='upload', name_columns_by_row=0)
#      # record_dict = request.get_records(field_name='upload')
#      month_year = datetime.now().strftime('%b-%Y')
#      month_year_folder = app.config["VANTAGE_UPLOAD_FOLDER"] + "/" + month_year
#
#      filename = secure_filename(request.files['upload'].filename)
#
#      filename_postfix_xlsx = Check_File_Exist(month_year_folder, ".".join(
#          filename.split(".")[:-1]) + ".xlsx")  # Checks, Creates folders and return AVAILABLE filename
#
#      # Want to Let the users download the File..
#      # return excel.make_response_from_records(record_dict, "xls", status=200, file_name=filename_without_postfix)
#
#      # pyexcel.save_as(records=record_dict, dest_file_name=filename_postfix_xlsx)
#
#      column_name = []
#      file_data = []
#      for cc, record in enumerate(record_dict):
#          if cc == 0:
#              column_name = list(record_dict[cc].keys())
#          buffer = dict()
#          # print(record)
#          for i, j in record.items():
#              if i == "":
#                  i = "Empty"
#              buffer[str(i).strip()] = str(j).strip()
#              print(i, j)
#          file_data.append(buffer)
#
#      raw_result = db.engine.execute("\
#          select `BGI_CORE`.`BGI_CORE_SYMBOL`,`BGI_VANTAGE`.`VANTAGE_CORE_SYMBOL`,`BGI_CORE`.`BGI_TYPE` , \
#          `BGI_CORE`.`BGI_CONTRACT_SIZE` , `BGI_CORE`.`BGI_DIGITS` , `VANTAGE_CORE`.`VANTAGE_DIGITS`,`VANTAGE_CORE`.`VANTAGE_CONTRACT_SIZE`, \
#          `BGI_CORE`.`BGI_POSITIVE_MARKUP`, `BGI_CORE`.`BGI_NEGATIVE_MARKUP`, `BGI_CORE`.currency, \
#          `BGI_FORCED`.`FORCED_BGI_LONG`, `BGI_FORCED`.`FORCED_BGI_SHORT`,  \
#          `BGI_FORCED_INSTI`.`BGI_INSTI_FORCED_LONG`,`BGI_FORCED_INSTI`.`BGI_INSTI_FORCED_SHORT` \
#          from \
#          (Select core_symbol as `BGI_CORE_SYMBOL`, contract_size as `BGI_CONTRACT_SIZE`, digits as `BGI_DIGITS`,  \
#          type as `BGI_TYPE`, positive_markup as `BGI_POSITIVE_MARKUP`, negative_markup as `BGI_NEGATIVE_MARKUP` ,currency \
#          from swap_bgicoresymbol) as `BGI_CORE` \
#          LEFT JOIN \
#          (Select bgi_coresymbol as `BGI_CORESYMBOL`, vantage_coresymbol as `VANTAGE_CORE_SYMBOL`  \
#          from aaron.swap_bgi_vantage_coresymbol) as `BGI_VANTAGE` on BGI_CORE.BGI_CORE_SYMBOL = BGI_VANTAGE.BGI_CORESYMBOL \
#          LEFT JOIN \
#          (Select core_symbol as `VANTAGE_CORESYMBOL`,contract_size as `VANTAGE_CONTRACT_SIZE`, digits as `VANTAGE_DIGITS`  \
#          from aaron.swap_vantagecoresymbol) as `VANTAGE_CORE` on `VANTAGE_CORE`.`VANTAGE_CORESYMBOL` = `BGI_VANTAGE`.`VANTAGE_CORE_SYMBOL` \
#          LEFT JOIN \
#          (Select core_symbol as `FORCED_BGI_CORESYMBOL`,`long` as `FORCED_BGI_LONG`, short as `FORCED_BGI_SHORT`  \
#          from aaron.swap_bgiforcedswap) as `BGI_FORCED` on BGI_CORE.BGI_CORE_SYMBOL = BGI_FORCED.FORCED_BGI_CORESYMBOL \
#          LEFT JOIN \
#          (Select core_symbol as `BGI_INSTI_FORCED_CORESYMBOL`, `long` as `BGI_INSTI_FORCED_LONG`, short as `BGI_INSTI_FORCED_SHORT`  \
#          from aaron.swap_bgiforcedswap_insti) as `BGI_FORCED_INSTI` on `BGI_CORE`.BGI_CORE_SYMBOL = `BGI_FORCED_INSTI`.BGI_INSTI_FORCED_CORESYMBOL \
#          order by FIELD(`BGI_CORE`.BGI_TYPE, 'FX','Exotic Pairs', 'PM', 'CFD'), BGI_CORE.BGI_CORE_SYMBOL \
#      ")
#
#      result_data = raw_result.fetchall()
#      result_col = raw_result.keys()
#      result_col_no_duplicate = []
#      for a in result_col:
#          if not a in result_col_no_duplicate:
#              result_col_no_duplicate.append(a)
#          else:
#              result_col_no_duplicate.append(str(a)+"_1")
#
#      collate = [dict(zip(result_col_no_duplicate, a)) for a in result_data]
#
#      # Calculate the Markup, as well as acount for the difference in Digits.
#      for i, c in enumerate(collate):     # By Per Row
#          collate[i]["SWAP_UPLOAD_LONG"] = ""         # We want to add the following..
#          collate[i]["SWAP_UPLOAD_LONG_MARKUP"] = ""
#          collate[i]["SWAP_UPLOAD_SHORT"] = ""
#          collate[i]["SWAP_UPLOAD_SHORT_MARKUP"] = ""
#
#          bgi_digit_difference = 0        # Sets as 0 for default.
#          if ("BGI_DIGITS"  in collate[i]) and ("VANTAGE_DIGITS" in collate[i]) and (collate[i]["BGI_DIGITS"] != None) and (collate[i]["VANTAGE_DIGITS"] != None):      # Need to calculate the Digit Difference.
#              bgi_digit_difference = int(collate[i]["BGI_DIGITS"]) - int(collate[i]["VANTAGE_DIGITS"])
#
#
#
#          # print(str(collate[i]['VANTAGE_CORE_SYMBOL']))
#
#          # Retail_Sheet.Cells(i, 2).Value = Find_Retail(Core_Symbol_BGI, 2, Cell_Color) * (10 ^ (Symbol_BGI_Digits - Symbol_Vantage_Digits))
#
#
#
#          for j, d in enumerate(file_data):
#              if 'VANTAGE_CORE_SYMBOL' in collate[i] and \
#                      len(list(d.keys())) >= 1 and \
#                      str(collate[i]['VANTAGE_CORE_SYMBOL']).strip() == d[list(d.keys())[0]]:
#
#                  d_key = [str(a).strip() for a in list(d.keys())]        # Get the keys for the Uploaded data.
#
#
#
#                  for ij, coll in enumerate(d_key):
#                      if "buy" in str(coll).lower() or "long" in str(coll).lower():   # Search for buy or long
#                          # Need to check if can be float.
#                          collate_keys = list(collate[i].keys())
#                          if "BGI_POSITIVE_MARKUP" in collate_keys and "BGI_NEGATIVE_MARKUP" in collate_keys and Check_Float(d[coll]):         # If posive, markup with positive markup. if negative, use negative markup.
#                              val = str(round(markup_swaps(float(str(d[coll])), collate[i]["BGI_POSITIVE_MARKUP"], collate[i]["BGI_NEGATIVE_MARKUP"])  * (10 ** bgi_digit_difference),4))  # Does the Digit Conversion here.
#                          else:
#                              val = ""
#                              flash("Unable to calculate markup prices for {}".format(collate[i]["BGI_CORE_SYMBOL"]))     # Put a message out that there is some error.
#
#                          collate[i]["SWAP_UPLOAD_LONG"] = str(d[coll])
#                          collate[i]["SWAP_UPLOAD_LONG_MARKUP"] = val
#
#
#                      if "sell" in str(coll).lower() or "short" in str(coll).lower(): # Search for sell or short in the colum names.
#
#                          # Need to check if can be float.
#                          collate_keys = list(collate[i].keys())
#                          if "BGI_POSITIVE_MARKUP" in collate_keys and "BGI_NEGATIVE_MARKUP" in collate_keys and Check_Float(d[coll]):    # If posive, markup with positive markup. if negative, use negative markup.
#                              val = str( round(markup_swaps(float(str(d[coll])), collate[i]["BGI_POSITIVE_MARKUP"], collate[i]["BGI_NEGATIVE_MARKUP"]) * (10 ** bgi_digit_difference)  ,4) ) # Does the Digit Conversion here.
#                          else:
#                              val = ""
#
#                          collate[i]["SWAP_UPLOAD_SHORT"] = str(d[coll])
#                          collate[i]["SWAP_UPLOAD_SHORT_MARKUP"] = val
#                  # print(d_key[0])
#                  break
#
#
#      # To Create the BGI Swaps.
#      bgi_long_swap_column = ["FORCED_BGI_LONG", "SWAP_UPLOAD_LONG_MARKUP"]       # To get the swap values in that order.
#      bgi_short_swap_column = ["FORCED_BGI_SHORT", "SWAP_UPLOAD_SHORT_MARKUP"]    # If there are forced swaps, get forced first.
#
#      bgi_insti_long_swap_column = ["FORCED_BGI_LONG", "BGI_INSTI_FORCED_LONG", "SWAP_UPLOAD_LONG_MARKUP"]    # If there are forced, if there are any Insti Forced, then the markup values.
#      bgi_insti_short_swap_column = ["FORCED_BGI_SHORT", "BGI_INSTI_FORCED_SHORT", "SWAP_UPLOAD_SHORT_MARKUP"]
#
#
#      bgi_swaps_retail = []
#      bgi_swaps_insti = []
#      for c in collate:
#          buffer_retail = {"SYMBOL": c["BGI_CORE_SYMBOL"], "LONG" : "", "SHORT" : ""}
#          buffer_insti ={"SYMBOL": c["BGI_CORE_SYMBOL"], "LONG" : "", "SHORT" : ""}
#
#          for l in bgi_long_swap_column:      # Going by Precedence.
#              if l in c and c[l] != None:
#                  buffer_retail["LONG"] = c[l]
#                  break
#          for s in bgi_short_swap_column:      # Going by Precedence.
#              if s in c and c[s] != None:
#                  buffer_retail["SHORT"] = c[s]
#                  break
#          for li in bgi_insti_long_swap_column:      # Going by Precedence.
#              if li in c and c[li] != None:
#                  buffer_insti["LONG"] = c[li]
#                  break
#          for si in bgi_insti_short_swap_column:      # Going by Precedence.
#              if si in c and c[si] != None:
#                  buffer_insti["SHORT"] = c[si]
#                  break
#
#          if buffer_retail["LONG"] == "" or buffer_retail["SHORT"] == "" or buffer_insti["LONG"] == "" or buffer_insti["SHORT"] == "" :
#              flash("{} has no Swaps.".format(c["BGI_CORE_SYMBOL"]))  # Flash out if there are no swaps found.
#
#          bgi_swaps_retail.append(buffer_retail)
#          bgi_swaps_insti.append(buffer_insti)
#
#
#      collate_table = [dict(zip([str(a).replace("_", " ") for a in c.keys()], c.values())) \
#                       for c in collate]
#
#
#      table = create_table_fun(collate_table)
#
#      table_bgi_swaps_retail = create_table_fun(bgi_swaps_retail)
#      table_bgi_swaps_insti = create_table_fun(bgi_swaps_insti)
#
#
#
#      return render_template("upload_form.html", table=table, form=form, table_bgi_swaps_retail=table_bgi_swaps_retail, table_bgi_swaps_insti=table_bgi_swaps_insti)
#
#  return render_template("upload_form.html", form=form)


# @app.route('/upload_swap3', methods=['GET', 'POST'])
# def retrieve_db_swaps3():
#  form = UploadForm()
#
#  if request.method == 'POST' and form.validate_on_submit():
#
#      # Get the file details as Records.
#      record_dict = request.get_records(field_name='upload', name_columns_by_row=0)
#
#      # ------------------ Dataframe of the data from CSV File. -----------------
#      df_csv = pd.DataFrame(record_dict)
#
#      uploaded_excel_data = df_csv.fillna("-").rename(columns=dict(zip(df_csv.columns,[a.replace("_","\n") for a in df_csv.columns]))).to_html(classes="table table-striped table-bordered table-hover table-condensed", index=False)
#
#
#      for i in df_csv.columns:  # Want to see which is Long and which is Short.
#          if "core symbol" in i.lower():
#              df_csv.rename(columns={i: "VANTAGE_CORE_SYMBOL"}, inplace=True) # Need to rename "Core symbol" to note that its from vantage.
#          if "long" in i.lower():
#              df_csv.rename(columns={i: "CSV_LONG"}, inplace=True)
#          if "short" in i.lower():
#              df_csv.rename(columns={i: "CSV_SHORT"}, inplace=True)
#
#
#      raw_result = db.engine.execute("\
#          select `BGI_CORE`.`BGI_CORE_SYMBOL`,`BGI_VANTAGE`.`VANTAGE_CORE_SYMBOL`,`BGI_CORE`.`BGI_TYPE` , \
#          `BGI_CORE`.`BGI_CONTRACT_SIZE` , `BGI_CORE`.`BGI_DIGITS` , `VANTAGE_CORE`.`VANTAGE_DIGITS`,`VANTAGE_CORE`.`VANTAGE_CONTRACT_SIZE`, \
#          `BGI_CORE`.`BGI_POSITIVE_MARKUP`, `BGI_CORE`.`BGI_NEGATIVE_MARKUP`, `BGI_CORE`.currency, \
#          `BGI_FORCED`.`FORCED_BGI_LONG`, `BGI_FORCED`.`FORCED_BGI_SHORT`,  \
#          `BGI_FORCED_INSTI`.`BGI_INSTI_FORCED_LONG`,`BGI_FORCED_INSTI`.`BGI_INSTI_FORCED_SHORT` \
#          from \
#          (Select core_symbol as `BGI_CORE_SYMBOL`, contract_size as `BGI_CONTRACT_SIZE`, digits as `BGI_DIGITS`,  \
#          type as `BGI_TYPE`, positive_markup as `BGI_POSITIVE_MARKUP`, negative_markup as `BGI_NEGATIVE_MARKUP` ,currency \
#          from swap_bgicoresymbol) as `BGI_CORE` \
#          LEFT JOIN \
#          (Select bgi_coresymbol as `BGI_CORESYMBOL`, vantage_coresymbol as `VANTAGE_CORE_SYMBOL`  \
#          from aaron.swap_bgi_vantage_coresymbol) as `BGI_VANTAGE` on BGI_CORE.BGI_CORE_SYMBOL = BGI_VANTAGE.BGI_CORESYMBOL \
#          LEFT JOIN \
#          (Select core_symbol as `VANTAGE_CORESYMBOL`,contract_size as `VANTAGE_CONTRACT_SIZE`, digits as `VANTAGE_DIGITS`  \
#          from aaron.swap_vantagecoresymbol) as `VANTAGE_CORE` on `VANTAGE_CORE`.`VANTAGE_CORESYMBOL` = `BGI_VANTAGE`.`VANTAGE_CORE_SYMBOL` \
#          LEFT JOIN \
#          (Select core_symbol as `FORCED_BGI_CORESYMBOL`,`long` as `FORCED_BGI_LONG`, short as `FORCED_BGI_SHORT`  \
#          from aaron.swap_bgiforcedswap) as `BGI_FORCED` on BGI_CORE.BGI_CORE_SYMBOL = BGI_FORCED.FORCED_BGI_CORESYMBOL \
#          LEFT JOIN \
#          (Select core_symbol as `BGI_INSTI_FORCED_CORESYMBOL`, `long` as `BGI_INSTI_FORCED_LONG`, short as `BGI_INSTI_FORCED_SHORT`  \
#          from aaron.swap_bgiforcedswap_insti) as `BGI_FORCED_INSTI` on `BGI_CORE`.BGI_CORE_SYMBOL = `BGI_FORCED_INSTI`.BGI_INSTI_FORCED_CORESYMBOL \
#          order by FIELD(`BGI_CORE`.BGI_TYPE, 'FX','Exotic Pairs', 'PM', 'CFD'), BGI_CORE.BGI_CORE_SYMBOL \
#      ")
#
#      result_data = raw_result.fetchall()
#      result_col = raw_result.keys()
#      result_col_no_duplicate = []
#      for a in result_col:
#          if not a in result_col_no_duplicate:
#              result_col_no_duplicate.append(a)
#          else:
#              result_col_no_duplicate.append(str(a)+"_1")
#
#      # Pandas data frame for the SQL return for the Symbol details.
#      df_sym_details = pd.DataFrame(data=result_data, columns=result_col_no_duplicate)
#      df_sym_details["DIGIT_DIFFERENCE"] = df_sym_details['BGI_DIGITS'] - df_sym_details['VANTAGE_DIGITS']    # Want to calculate the Digit Difference.
#
#
#      # ---------------------------- Time to merge the 2 Data Frames. ---------------------------------------------
#      combine_df = df_sym_details.merge(df_csv, on=["VANTAGE_CORE_SYMBOL"], how="outer")
#      combine_df = combine_df[pd.notnull(combine_df["BGI_CORE_SYMBOL"])]
#
#      # Need to Flip LONG
#      combine_df["CSV_LONG"] = combine_df[ "CSV_LONG"] * -1  # Vantage sending us Positive = Charged. We need to flip  LONG
#      # Need to correct the number of digits.
#      if "CSV_LONG" in combine_df.columns:
#          combine_df["CSV_LONG_CORRECT_DIGITS"] = combine_df["CSV_LONG"] * (10 ** combine_df["DIGIT_DIFFERENCE"])
#      if "CSV_SHORT" in combine_df.columns:
#          combine_df["CSV_SHORT_CORRECT_DIGITS"] = combine_df["CSV_SHORT"] * (10 ** combine_df["DIGIT_DIFFERENCE"])
#
#      # Want to multiply by the correct markup for each. Want to give less, take more.
#      # Long
#      combine_df["CSV_LONG_CORRECT_DIGITS_MARKUP"] = np.where(combine_df["CSV_LONG_CORRECT_DIGITS"] > 0, round(
#          combine_df["CSV_LONG_CORRECT_DIGITS"] * (1 - (combine_df["BGI_POSITIVE_MARKUP"] / 100)), 3), round(
#          combine_df["CSV_LONG_CORRECT_DIGITS"] * (1 + (combine_df["BGI_NEGATIVE_MARKUP"] / 100)), 3))
#      # Short
#      combine_df["CSV_SHORT_CORRECT_DIGITS_MARKUP"] = np.where(combine_df["CSV_SHORT_CORRECT_DIGITS"] > 0, round(
#          combine_df["CSV_SHORT_CORRECT_DIGITS"] * (1 - (combine_df["BGI_POSITIVE_MARKUP"] / 100)), 3), round(
#          combine_df["CSV_SHORT_CORRECT_DIGITS"] * (1 + (combine_df["BGI_NEGATIVE_MARKUP"] / 100)), 3))
#
#      # The Data Frame used for the Building of the CSV File.
#      build_bgi_swap_df = combine_df[
#          ["BGI_CORE_SYMBOL", "VANTAGE_CORE_SYMBOL", "BGI_POSITIVE_MARKUP", "BGI_NEGATIVE_MARKUP", "FORCED_BGI_LONG",
#           "FORCED_BGI_SHORT", "BGI_INSTI_FORCED_LONG", "BGI_INSTI_FORCED_SHORT", "CSV_LONG_CORRECT_DIGITS_MARKUP",
#           "CSV_SHORT_CORRECT_DIGITS_MARKUP", "BGI_TYPE"]]
#
#      # Want to get either the Forced Swaps, if not, Get the "Vantage corrected digit markup" swaps
#      build_bgi_swap_df.loc[:, "LONG"] = np.where(pd.notnull(build_bgi_swap_df["FORCED_BGI_LONG"]),
#                                                  round(build_bgi_swap_df["FORCED_BGI_LONG"], 3), np.where(
#              pd.notnull(build_bgi_swap_df["CSV_LONG_CORRECT_DIGITS_MARKUP"]),
#              round(build_bgi_swap_df["CSV_LONG_CORRECT_DIGITS_MARKUP"], 3),
#              build_bgi_swap_df["CSV_LONG_CORRECT_DIGITS_MARKUP"]))
#      build_bgi_swap_df.loc[:, "SHORT"] = np.where(pd.notnull(build_bgi_swap_df["FORCED_BGI_SHORT"]),
#                                                   round(build_bgi_swap_df["FORCED_BGI_SHORT"], 3), np.where(
#              pd.notnull(build_bgi_swap_df["CSV_SHORT_CORRECT_DIGITS_MARKUP"]),
#              round(build_bgi_swap_df["CSV_SHORT_CORRECT_DIGITS_MARKUP"], 3),
#              build_bgi_swap_df["CSV_SHORT_CORRECT_DIGITS_MARKUP"]))
#
#      # For Insti, we want to check BGI_FORCED first, if null, check BGI_INSTI_FORCED. If not, use vantage digit change markup swap.
#      build_bgi_swap_df.loc[:, "INSTI_LONG"] = np.where(pd.notnull(build_bgi_swap_df["FORCED_BGI_LONG"]), round(build_bgi_swap_df["FORCED_BGI_LONG"], 3), \
#                                                        np.where( pd.notnull(build_bgi_swap_df["BGI_INSTI_FORCED_LONG"]), round(build_bgi_swap_df["BGI_INSTI_FORCED_LONG"], 3), \
#                                                            np.where(pd.notnull( build_bgi_swap_df["CSV_LONG_CORRECT_DIGITS_MARKUP"]), round(build_bgi_swap_df[ "CSV_LONG_CORRECT_DIGITS_MARKUP"], 3), \
#                                                                     build_bgi_swap_df["CSV_LONG_CORRECT_DIGITS_MARKUP"])))
#
#
#      build_bgi_swap_df.loc[:, "INSTI_SHORT"] = np.where(pd.notnull(build_bgi_swap_df["FORCED_BGI_SHORT"]), round(build_bgi_swap_df["FORCED_BGI_SHORT"], 3), \
#                                                         np.where( pd.notnull(build_bgi_swap_df["BGI_INSTI_FORCED_SHORT"]), round(build_bgi_swap_df["BGI_INSTI_FORCED_SHORT"], 3), \
#                                                             np.where(pd.notnull(build_bgi_swap_df["CSV_SHORT_CORRECT_DIGITS_MARKUP"]), round(build_bgi_swap_df["CSV_SHORT_CORRECT_DIGITS_MARKUP"], 3), \
#                                                                      build_bgi_swap_df["CSV_SHORT_CORRECT_DIGITS_MARKUP"])))
#
#      # Want to find out which of the symbols still have NULL.
#
#      Swap_Error = build_bgi_swap_df[pd.isnull(build_bgi_swap_df["LONG"]) | pd.isnull(build_bgi_swap_df["SHORT"])]
#      if len(Swap_Error) != 0:
#          for a in Swap_Error["BGI_CORE_SYMBOL"]:
#              flash("{} has swap errors. Swap is null.".format(a))
#
#      # Minimising the data frame. Want to compare with median of last 15 days.
#      build_bgi_swap_df_show = build_bgi_swap_df[["BGI_CORE_SYMBOL", "LONG", "SHORT", "INSTI_LONG", "INSTI_SHORT"]]   # Data Frame
#
#      raw_result2 = db.engine.execute("SELECT * FROM test.bgi_swaps where date > CURDATE()-15")
#
#      result_data2 = raw_result2.fetchall()
#      result_col2 = raw_result2.keys()
#
#
#      cfd_swaps = pd.DataFrame(data=result_data2, columns=result_col2)
#      cfd_swaps.loc[:, "bgi_long"] = pd.to_numeric(cfd_swaps["bgi_long"])
#      cfd_swaps.loc[:, "bgi_short"] = pd.to_numeric(cfd_swaps["bgi_short"])
#      cfd_swaps.rename(columns={"Core_Symbol": "BGI_CORE_SYMBOL", "bgi_long":"BGI_LONG_AVERAGE", "bgi_short":"BGI_SHORT_AVERAGE"}, inplace=True)  # Rename for easy join.
#      cfd_swaps_median = cfd_swaps.groupby("BGI_CORE_SYMBOL").median()
#      combine_average_df = build_bgi_swap_df_show.join(cfd_swaps_median, on=["BGI_CORE_SYMBOL"])
#
#      # combine_average_df[]
#
#      # Data that we want to show.
#
#      # table_bgi_swaps_retail = combine_average_df.fillna("-").rename(columns=dict(zip(combine_average_df.columns,[a.replace("_","\n") for a in combine_average_df.columns]))).to_html(classes="table table-striped table-bordered table-hover table-condensed", index=False)
#      table_bgi_swaps_retail = combine_average_df.fillna("-").rename(columns=dict(zip(combine_average_df.columns,[a.replace("_","\n") for a in combine_average_df.columns]))).style.hide_index().applymap(color_negative_red, subset=["LONG","SHORT"]).set_table_attributes('class="table table-striped table-bordered table-hover table-condensed"').render()
#
#      table_bgi_swaps_insti = ""
#
#
#      table = combine_df.fillna("-").rename(columns=dict(zip(combine_df.columns,[a.replace("_","\n") for a in combine_df.columns]))).to_html(classes="table table-striped table-bordered table-hover table-condensed", index=False)
#
#      return render_template("upload_form2.html", table=table, form=form, table_bgi_swaps_retail=table_bgi_swaps_retail, table_bgi_swaps_insti=table_bgi_swaps_insti, uploaded_excel_data = uploaded_excel_data)
#
#  return render_template("upload_form.html", form=form)
#
